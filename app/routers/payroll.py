"""Payroll router - Prima nota and salary management."""
import io
from openpyxl import load_workbook
import re
from fastapi import Body, HTTPException

from fastapi import APIRouter, Depends, Path, status, UploadFile, File, Query
from typing import Dict, Any, List
from datetime import datetime, timezone
from uuid import uuid4
import logging

from app.database import Database
from app.utils.dependencies import get_current_user

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get(
    "/prima-nota",
    summary="Get prima nota entries"
)
async def get_prima_nota(
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> List[Dict[str, Any]]:
    """Get prima nota entries."""
    db = Database.get_db()
    entries = await db["prima_nota_cassa"].find({}, {"_id": 0}).sort("date", -1).to_list(500)
    return entries


@router.post(
    "/prima-nota",
    status_code=status.HTTP_201_CREATED,
    summary="Create prima nota entry"
)
async def create_prima_nota(
    data: Dict[str, Any] = Body(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, str]:
    """Create a prima nota entry."""
    db = Database.get_db()
    data["id"] = str(uuid4())
    data["created_at"] = datetime.now(timezone.utc)
    await db["prima_nota_cassa"].insert_one(data.copy())
    return {"message": "Entry created", "id": data["id"]}


@router.put(
    "/prima-nota/{entry_id}/note",
    summary="Update entry note"
)
async def update_note(
    entry_id: str = Path(...),
    note: str = "",
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, str]:
    """Update prima nota entry note."""
    db = Database.get_db()
    await db["prima_nota_cassa"].update_one({"id": entry_id}, {"$set": {"note": note}})
    return {"message": "Note updated"}


@router.delete(
    "/prima-nota/delete-all",
    summary="Delete all prima nota"
)
async def delete_all_prima_nota(
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, str]:
    """Delete all prima nota entries."""
    db = Database.get_db()
    result = await db["prima_nota_cassa"].delete_many({})
    return {"message": f"Deleted {result.deleted_count} entries"}


@router.post(
    "/prima-nota/import-excel",
    summary="Import from Excel"
)
async def import_excel(
    file: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """Import prima nota from Excel."""
    db = Database.get_db()
    
    # Check extension
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format")
    
    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    
    # Select best sheet
    if "Dettaglio" in wb.sheetnames:
        ws = wb["Dettaglio"]
    elif "Riepilogo" in wb.sheetnames:
        ws = wb["Riepilogo"]
    else:
        ws = wb.active
        
    # Extract year from filename (default)
    default_year = datetime.now(timezone.utc).year
    date_match = re.search(r'(\d{4})', file.filename)
    if date_match:
        default_year = int(date_match.group(1))
        
    headers = {}
    row_count = 0
    imported_count = 0
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # Parse headers
            for idx, cell in enumerate(row):
                if cell:
                    headers[str(cell).upper().strip()] = idx
            continue
            
        # Determine columns based on headers found
        dipendente = None
        
        # Try finding Employee column
        for col_name in ['DIPENDENTE', 'NOME', 'COGNOME NOME', 'DIPENDENTE/COLLABORATORE']:
            if col_name in headers:
                dipendente = row[headers[col_name]]
                break
        
        if not dipendente:
            continue
            
        # Initialize record values
        stipendio_netto = 0.0
        importo_erogato = 0.0
        saldo = 0.0
        anno = default_year
        mese = "Riepilogo"
        
        # Mapping Logic
        
        # 1. Netto / Busta
        for col_name in ['IMPORTO BUSTA', 'TOTALE BUSTE', 'NETTO', 'STIPENDIO', 'NETTO BUSTA']:
            if col_name in headers:
                val = row[headers[col_name]]
                if val:
                    if isinstance(val, str):
                        val = val.replace('.', '').replace(',', '.')
                    stipendio_netto = float(val)
                break
                
        # 2. Erogato / Pagato
        for col_name in ['IMPORTO PAGATO', 'TOTALE PAGATO', 'EROGATO', 'PAGATO']:
            if col_name in headers:
                val = row[headers[col_name]]
                if val:
                    if isinstance(val, str):
                        val = val.replace('.', '').replace(',', '.')
                    importo_erogato = float(val)
                break
                
        # 3. Saldo
        for col_name in ['SALDO', 'DIFFERENZA']:
            if col_name in headers:
                val = row[headers[col_name]]
                if val:
                    if isinstance(val, str):
                        val = val.replace('.', '').replace(',', '.')
                    saldo = float(val)
                break
                
        # 4. Anno / Mese (from Row if available)
        if 'ANNO' in headers:
            val = row[headers['ANNO']]
            if val:
                anno = int(val)
                
        if 'MESE' in headers:
            val = row[headers['MESE']]
            if val:
                mese = str(val).capitalize()
        
        # If saldo not explicit, calc it
        if 'SALDO' not in headers and stipendio_netto and importo_erogato:
            saldo = stipendio_netto - importo_erogato
        
        # Create record
        doc = {
            "id": str(uuid4()),
            "user_id": current_user["user_id"],
            "created_at": datetime.now(timezone.utc),
            "anno": anno,
            "mese": mese,
            "dipendente_nome": dipendente,
            "stipendio_netto": stipendio_netto,
            "importo_erogato": importo_erogato,
            "saldo": saldo,
            "saldo_progressivo": saldo, # Init
            "note": f"Importato da {file.filename}",
            "source": "excel_import"
        }
        
        await db["prima_nota_cassa"].insert_one(doc.copy())
        imported_count += 1
        
    return {
        "message": "Import completed", 
        "imported": imported_count,
        "year": default_year
    }


@router.post(
    "/prima-nota/import-bonifici-excel",
    summary="Import bank transfers from Excel"
)
async def import_bonifici(
    file: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """Import bank transfers from Excel."""
    return {"message": "Import completed", "imported": 0}


@router.get(
    "/prima-nota/dipendenti",
    summary="Get employees for payroll"
)
async def get_prima_nota_dipendenti(
    anno: int = Query(None),
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> List[Dict[str, Any]]:
    """Get employees for payroll prima nota."""
    db = Database.get_db()
    from app.database import Collections
    employees = await db[Collections.EMPLOYEES].find({}, {"_id": 0}).to_list(500)
    return employees