"""
Router per gestione invio documenti al Commercialista.

Permette l'invio via email di:
- Prima Nota Cassa mensile (PDF)
- Carnet assegni (PDF)
- Fatture pagate per cassa (PDF)
"""
from fastapi import APIRouter, HTTPException, Body
from typing import Dict, Any, Optional
from datetime import datetime, timezone
from calendar import monthrange
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os
import base64

from app.database import Database
from app.utils.error_handler import handle_errors

logger = logging.getLogger(__name__)
router = APIRouter()

# Email commercialista di default
DEFAULT_COMMERCIALISTA_EMAIL = "rosaria.marotta@email.it"


def get_smtp_config():
    """Get SMTP configuration from environment."""
    return {
        "host": os.environ.get('SMTP_HOST', 'smtp.gmail.com'),
        "port": int(os.environ.get('SMTP_PORT', 587)),
        "user": (
            os.environ.get('SMTP_USER')
            or os.environ.get('SMTP_USERNAME')
            or os.environ.get('EMAIL_USER')
        ),
        "password": (
            os.environ.get('SMTP_PASSWORD')
            or os.environ.get('EMAIL_PASSWORD')
        ),
        "from_email": (
            os.environ.get('FROM_EMAIL')
            or os.environ.get('SMTP_FROM_EMAIL')
            or os.environ.get('SMTP_USER')
            or os.environ.get('EMAIL_USER')
        ),
    }


def send_email_with_attachment(
    to_email: str, 
    subject: str, 
    html_body: str, 
    attachment_data: Optional[bytes] = None,
    attachment_name: Optional[str] = None
) -> bool:
    """Send email with optional PDF attachment."""
    config = get_smtp_config()
    
    if not all([config["host"], config["user"], config["password"]]):
        logger.error("SMTP not configured")
        raise HTTPException(status_code=500, detail="Configurazione SMTP mancante")
    
    try:
        msg = MIMEMultipart()
        msg['Subject'] = subject
        msg['From'] = config["from_email"] or config["user"]
        msg['To'] = to_email
        
        # HTML body
        html_part = MIMEText(html_body, 'html', 'utf-8')
        msg.attach(html_part)
        
        # Attachment if provided
        if attachment_data and attachment_name:
            part = MIMEBase('application', 'pdf')
            part.set_payload(attachment_data)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{attachment_name}"')
            msg.attach(part)
        
        with smtplib.SMTP(config["host"], config["port"], timeout=30) as server:
            server.starttls()
            server.login(config["user"], config["password"])
            server.send_message(msg)
        
        logger.info(f"Email sent to {to_email}: {subject}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send email to {to_email}: {e}")
        raise HTTPException(status_code=500, detail=f"Errore invio email: {str(e)}")


@router.get("/config")
@handle_errors
async def get_commercialista_config() -> Dict[str, Any]:
    """Get commercialista configuration."""
    db = Database.get_db()
    
    # Try to get config from DB
    config = await db["commercialista_config"].find_one({}, {"_id": 0})
    
    if not config:
        config = {
            "email": DEFAULT_COMMERCIALISTA_EMAIL,
            "nome": "Dott.ssa Rosaria Marotta",
            "alert_giorni": 2,
            "invio_automatico": False
        }
    
    # Add SMTP status
    smtp_config = get_smtp_config()
    config["smtp_configured"] = all([smtp_config["host"], smtp_config["user"], smtp_config["password"]])
    
    return config


@router.put("/config")
@handle_errors
async def update_commercialista_config(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Update commercialista configuration."""
    db = Database.get_db()
    
    update_data = {
        "email": data.get("email", DEFAULT_COMMERCIALISTA_EMAIL),
        "nome": data.get("nome", ""),
        "alert_giorni": data.get("alert_giorni", 2),
        "invio_automatico": data.get("invio_automatico", False),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db["commercialista_config"].update_one(
        {},
        {"$set": update_data},
        upsert=True
    )
    
    return {"success": True, "config": update_data}


@router.get("/prima-nota-cassa/{anno}/{mese}")
@handle_errors
async def get_prima_nota_cassa_mensile(anno: int, mese: int) -> Dict[str, Any]:
    """Get Prima Nota Cassa for a specific month."""
    db = Database.get_db()
    
    # Date range for the month - usa formato YYYY-MM per regex match
    month_prefix = f"{anno}-{mese:02d}"
    _, last_day = monthrange(anno, mese)
    start_date = f"{anno}-{mese:02d}-01"
    end_date = f"{anno}-{mese:02d}-{last_day:02d}"
    
    # Query prima nota cassa - collection principale
    movements = []
    
    # Try prima_nota_cassa collection (usata dal router prima_nota.py)
    # Ordinamento: prima per data, poi per categoria (Corrispettivi prima di POS)
    cursor = db["prima_nota_cassa"].find({
        "data": {"$regex": f"^{month_prefix}"}
    }, {"_id": 0}).sort([("data", 1), ("categoria", 1)])
    movements = await cursor.to_list(5000)
    
    # If empty, try prima_nota collection with tipo_conto = cassa
    if not movements:
        cursor = db["prima_nota_cassa"].find({
            "tipo_conto": "cassa",
            "$or": [
                {"data": {"$regex": f"^{month_prefix}"}},
                {"date": {"$regex": f"^{month_prefix}"}}
            ]
        }, {"_id": 0}).sort([("data", 1), ("categoria", 1)])
        movements = await cursor.to_list(5000)
    
    # If still empty, try cash collection
    if not movements:
        cursor = db["cash"].find({
            "$or": [
                {"data": {"$regex": f"^{month_prefix}"}},
                {"date": {"$regex": f"^{month_prefix}"}}
            ]
        }, {"_id": 0}).sort([("data", 1), ("categoria", 1)])
        movements = await cursor.to_list(5000)
    
    # Calculate totals
    totale_entrate = 0
    totale_uscite = 0
    
    for m in movements:
        tipo = m.get("type") or m.get("tipo") or ""
        importo = float(m.get("amount") or m.get("importo") or 0)
        
        if tipo.lower() in ["entrata", "income", "in"]:
            totale_entrate += abs(importo)
        else:
            totale_uscite += abs(importo)
    
    saldo = totale_entrate - totale_uscite
    
    return {
        "anno": anno,
        "mese": mese,
        "mese_nome": ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
                     "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"][mese],
        "movimenti": movements,
        "totale_movimenti": len(movements),
        "totale_entrate": round(totale_entrate, 2),
        "totale_uscite": round(totale_uscite, 2),
        "saldo": round(saldo, 2)
    }


@router.get("/fatture-cassa/{anno}/{mese}")
@handle_errors
async def get_fatture_pagate_cassa(anno: int, mese: int) -> Dict[str, Any]:
    """Get invoices paid by cash for a specific month."""
    db = Database.get_db()
    
    month_prefix = f"{anno}-{mese:02d}"
    
    # Query fatture with payment method = contanti/cassa and date in month
    cursor = db["invoices"].find({
        "$and": [
            {"$or": [
                {"metodo_pagamento": {"$regex": "contant|cassa", "$options": "i"}},
                {"payment_method": {"$regex": "contant|cassa|cash", "$options": "i"}},
                {"modalita_pagamento": {"$regex": "contant|cassa", "$options": "i"}}
            ]},
            {"$or": [
                {"data_pagamento": {"$regex": f"^{month_prefix}"}},
                {"invoice_date": {"$regex": f"^{month_prefix}"}},
                {"data_fattura": {"$regex": f"^{month_prefix}"}}
            ]}
        ]
    }, {"_id": 0})
    
    fatture = await cursor.to_list(10000)
    totale = sum(float(f.get("total_amount") or f.get("importo_totale") or 0) for f in fatture)
    
    return {
        "anno": anno,
        "mese": mese,
        "mese_nome": ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
                     "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"][mese],
        "fatture": fatture,
        "totale_fatture": len(fatture),
        "totale_importo": round(totale, 2)
    }


@router.post("/invia-prima-nota")
@handle_errors
async def invia_prima_nota_cassa(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Send Prima Nota Cassa via email with PDF attachment."""
    anno = data.get("anno")
    mese = data.get("mese")
    email = data.get("email", DEFAULT_COMMERCIALISTA_EMAIL)
    pdf_base64 = data.get("pdf_base64")  # PDF generated by frontend
    
    if not anno or not mese:
        raise HTTPException(status_code=400, detail="Anno e mese richiesti")
    
    mese_nome = ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
                 "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"][mese]
    
    # Get data for email body
    prima_nota_data = await get_prima_nota_cassa_mensile(anno, mese)
    
    subject = f"📒 Prima Nota Cassa - {mese_nome} {anno}"
    
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; padding: 20px; background: #f5f5f5;">
        <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
            <div style="background: linear-gradient(135deg, #1e3a5f 0%, #2563eb 100%); color: white; padding: 20px;">
                <h1 style="margin: 0;">📒 Prima Nota Cassa</h1>
                <p style="margin: 10px 0 0 0; opacity: 0.9;">{mese_nome} {anno}</p>
            </div>
            
            <div style="padding: 20px;">
                <p>Gentile Commercialista,</p>
                <p>in allegato trova la Prima Nota Cassa relativa al mese di <strong>{mese_nome} {anno}</strong>.</p>
                
                <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <h3 style="margin: 0 0 10px 0; color: #1e3a5f;">📊 Riepilogo</h3>
                    <table style="width: 100%;">
                        <tr>
                            <td style="padding: 5px 0;">Movimenti totali:</td>
                            <td style="padding: 5px 0; text-align: right; font-weight: bold;">{prima_nota_data['totale_movimenti']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 5px 0; color: #4caf50;">Totale Entrate:</td>
                            <td style="padding: 5px 0; text-align: right; font-weight: bold; color: #4caf50;">€ {prima_nota_data['totale_entrate']:,.2f}</td>
                        </tr>
                        <tr>
                            <td style="padding: 5px 0; color: #f44336;">Totale Uscite:</td>
                            <td style="padding: 5px 0; text-align: right; font-weight: bold; color: #f44336;">€ {prima_nota_data['totale_uscite']:,.2f}</td>
                        </tr>
                        <tr style="border-top: 2px solid #ddd;">
                            <td style="padding: 10px 0 5px 0; font-weight: bold;">Saldo:</td>
                            <td style="padding: 10px 0 5px 0; text-align: right; font-weight: bold; font-size: 18px; color: {'#4caf50' if prima_nota_data['saldo'] >= 0 else '#f44336'};">€ {prima_nota_data['saldo']:,.2f}</td>
                        </tr>
                    </table>
                </div>
                
                <p style="color: #666; font-size: 14px;">
                    Il documento PDF allegato contiene il dettaglio completo di tutti i movimenti.
                </p>
            </div>
            
            <div style="background: #f5f5f5; padding: 15px; text-align: center; font-size: 12px; color: #666;">
                Ceraldi Group S.R.L. - ERP Azienda Semplice<br>
                Email generata automaticamente il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}
            </div>
        </div>
    </body>
    </html>
    """
    
    # Decode PDF if provided
    pdf_bytes = None
    if pdf_base64:
        try:
            pdf_bytes = base64.b64decode(pdf_base64)
        except Exception as e:
            logger.error(f"Error decoding PDF: {e}")
    
    filename = f"Prima_Nota_Cassa_{mese_nome}_{anno}.pdf"
    
    success = send_email_with_attachment(email, subject, html_body, pdf_bytes, filename)
    
    # Log the send
    db = Database.get_db()
    log_doc = {
        "tipo": "prima_nota_cassa",
        "anno": anno,
        "mese": mese,
        "email": email,
        "data_invio": datetime.now(timezone.utc).isoformat(),
        "success": success
    }
    await db["commercialista_log"].insert_one(log_doc.copy())
    
    return {
        "success": success,
        "message": f"Prima Nota Cassa {mese_nome} {anno} inviata a {email}"
    }


@router.post("/invia-carnet")
@handle_errors
async def invia_carnet(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Send Carnet assegni via email with PDF attachment."""
    carnet_id = data.get("carnet_id")
    email = data.get("email", DEFAULT_COMMERCIALISTA_EMAIL)
    pdf_base64 = data.get("pdf_base64")
    assegni_count = data.get("assegni_count", 0)
    totale_importo = data.get("totale_importo", 0)
    
    if not carnet_id:
        raise HTTPException(status_code=400, detail="carnet_id richiesto")
    
    subject = f"📝 Carnet Assegni - {carnet_id}"
    
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; padding: 20px; background: #f5f5f5;">
        <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
            <div style="background: linear-gradient(135deg, #4caf50 0%, #2e7d32 100%); color: white; padding: 20px;">
                <h1 style="margin: 0;">📝 Carnet Assegni</h1>
                <p style="margin: 10px 0 0 0; opacity: 0.9;">ID: {carnet_id}</p>
            </div>
            
            <div style="padding: 20px;">
                <p>Gentile Commercialista,</p>
                <p>in allegato trova il riepilogo del carnet assegni <strong>{carnet_id}</strong>.</p>
                
                <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <h3 style="margin: 0 0 10px 0; color: #2e7d32;">📊 Riepilogo Carnet</h3>
                    <table style="width: 100%;">
                        <tr>
                            <td style="padding: 5px 0;">Numero Assegni:</td>
                            <td style="padding: 5px 0; text-align: right; font-weight: bold;">{assegni_count}</td>
                        </tr>
                        <tr style="border-top: 1px solid #ddd;">
                            <td style="padding: 10px 0 5px 0; font-weight: bold;">Totale Importo:</td>
                            <td style="padding: 10px 0 5px 0; text-align: right; font-weight: bold; font-size: 18px; color: #2e7d32;">€ {totale_importo:,.2f}</td>
                        </tr>
                    </table>
                </div>
                
                <p style="color: #666; font-size: 14px;">
                    Il documento PDF allegato contiene il dettaglio di tutti gli assegni del carnet.
                </p>
            </div>
            
            <div style="background: #f5f5f5; padding: 15px; text-align: center; font-size: 12px; color: #666;">
                Ceraldi Group S.R.L. - ERP Azienda Semplice<br>
                Email generata automaticamente il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}
            </div>
        </div>
    </body>
    </html>
    """
    
    pdf_bytes = None
    if pdf_base64:
        try:
            pdf_bytes = base64.b64decode(pdf_base64)
        except Exception as e:
            logger.error(f"Error decoding PDF: {e}")
    
    filename = f"Carnet_Assegni_{carnet_id}.pdf"
    
    success = send_email_with_attachment(email, subject, html_body, pdf_bytes, filename)
    
    # Log the send
    db = Database.get_db()
    log_doc = {
        "tipo": "carnet_assegni",
        "carnet_id": carnet_id,
        "email": email,
        "data_invio": datetime.now(timezone.utc).isoformat(),
        "success": success
    }
    await db["commercialista_log"].insert_one(log_doc.copy())
    
    return {
        "success": success,
        "message": f"Carnet {carnet_id} inviato a {email}"
    }


@router.post("/invia-fatture-cassa")
@handle_errors
async def invia_fatture_cassa(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Send fatture pagate per cassa via email with PDF attachment."""
    anno = data.get("anno")
    mese = data.get("mese")
    email = data.get("email", DEFAULT_COMMERCIALISTA_EMAIL)
    pdf_base64 = data.get("pdf_base64")
    
    if not anno or not mese:
        raise HTTPException(status_code=400, detail="Anno e mese richiesti")
    
    mese_nome = ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
                 "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"][mese]
    
    # Get data for email body
    fatture_data = await get_fatture_pagate_cassa(anno, mese)
    
    subject = f"💵 Fatture Pagate per Cassa - {mese_nome} {anno}"
    
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; padding: 20px; background: #f5f5f5;">
        <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
            <div style="background: linear-gradient(135deg, #ff9800 0%, #f57c00 100%); color: white; padding: 20px;">
                <h1 style="margin: 0;">💵 Fatture Pagate per Cassa</h1>
                <p style="margin: 10px 0 0 0; opacity: 0.9;">{mese_nome} {anno}</p>
            </div>
            
            <div style="padding: 20px;">
                <p>Gentile Commercialista,</p>
                <p>in allegato trova l'elenco delle fatture pagate per cassa nel mese di <strong>{mese_nome} {anno}</strong>.</p>
                
                <div style="background: #fff3e0; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <h3 style="margin: 0 0 10px 0; color: #f57c00;">📊 Riepilogo</h3>
                    <table style="width: 100%;">
                        <tr>
                            <td style="padding: 5px 0;">Numero Fatture:</td>
                            <td style="padding: 5px 0; text-align: right; font-weight: bold;">{fatture_data['totale_fatture']}</td>
                        </tr>
                        <tr style="border-top: 1px solid #ffe0b2;">
                            <td style="padding: 10px 0 5px 0; font-weight: bold;">Totale:</td>
                            <td style="padding: 10px 0 5px 0; text-align: right; font-weight: bold; font-size: 18px; color: #f57c00;">€ {fatture_data['totale_importo']:,.2f}</td>
                        </tr>
                    </table>
                </div>
                
                <p style="color: #666; font-size: 14px;">
                    Il documento PDF allegato contiene il dettaglio di tutte le fatture.
                </p>
            </div>
            
            <div style="background: #f5f5f5; padding: 15px; text-align: center; font-size: 12px; color: #666;">
                Ceraldi Group S.R.L. - ERP Azienda Semplice<br>
                Email generata automaticamente il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}
            </div>
        </div>
    </body>
    </html>
    """
    
    pdf_bytes = None
    if pdf_base64:
        try:
            pdf_bytes = base64.b64decode(pdf_base64)
        except Exception as e:
            logger.error(f"Error decoding PDF: {e}")
    
    filename = f"Fatture_Contanti_{mese_nome}_{anno}.pdf"
    
    success = send_email_with_attachment(email, subject, html_body, pdf_bytes, filename)
    
    # Log the send
    db = Database.get_db()
    log_doc = {
        "tipo": "fatture_cassa",
        "anno": anno,
        "mese": mese,
        "email": email,
        "data_invio": datetime.now(timezone.utc).isoformat(),
        "success": success
    }
    await db["commercialista_log"].insert_one(log_doc.copy())
    
    return {
        "success": success,
        "message": f"Fatture contanti {mese_nome} {anno} inviate a {email}"
    }


@router.get("/log")
@handle_errors
async def get_invio_log(limit: int = 50) -> Dict[str, Any]:
    """Get log of sent documents."""
    db = Database.get_db()
    
    cursor = db["commercialista_log"].find({}, {"_id": 0}).sort("data_invio", -1).limit(limit)
    log_entries = await cursor.to_list(limit)
    
    return {
        "log": log_entries,
        "totale": len(log_entries)
    }


@router.post("/segna-inviata")
@handle_errors
async def segna_prima_nota_inviata(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Segna manualmente la Prima Nota Cassa come inviata per un determinato mese/anno."""
    anno = data.get("anno")
    mese = data.get("mese")
    email = data.get("email", DEFAULT_COMMERCIALISTA_EMAIL)
    
    if not anno or not mese:
        raise HTTPException(status_code=400, detail="anno e mese sono obbligatori")
    
    db = Database.get_db()
    log_doc = {
        "tipo": "prima_nota_cassa",
        "anno": anno,
        "mese": mese,
        "email": email,
        "data_invio": datetime.now(timezone.utc).isoformat(),
        "success": True,
        "note": "Segnata manualmente come inviata"
    }
    await db["commercialista_log"].insert_one(log_doc.copy())
    
    mese_nome = ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
                 "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"][mese]
    return {
        "success": True,
        "message": f"Prima Nota Cassa {mese_nome} {anno} segnata come inviata"
    }


@router.get("/alert-status")
@handle_errors
async def get_alert_status() -> Dict[str, Any]:
    """Check if there are pending documents to send (for alert)."""
    now = datetime.now(timezone.utc)
    current_month = now.month
    current_year = now.year
    
    # Previous month
    if current_month == 1:
        prev_month = 12
        prev_year = current_year - 1
    else:
        prev_month = current_month - 1
        prev_year = current_year
    
    # Check if we're within 2 days of the month end
    _, last_day = monthrange(prev_year, prev_month)
    deadline = datetime(current_year, current_month, 2, 23, 59, 59, tzinfo=timezone.utc)
    
    db = Database.get_db()
    
    # Check if prima nota was already sent for previous month
    prima_nota_sent = await db["commercialista_log"].find_one({
        "tipo": "prima_nota_cassa",
        "anno": prev_year,
        "mese": prev_month,
        "success": True
    })
    
    show_alert = now <= deadline and not prima_nota_sent
    
    mese_nome = ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
                 "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"][prev_month]
    
    return {
        "show_alert": show_alert,
        "mese_pendente": prev_month,
        "anno_pendente": prev_year,
        "mese_nome": mese_nome,
        "deadline": deadline.isoformat(),
        "prima_nota_inviata": prima_nota_sent is not None,
        "message": f"Ricordati di inviare la Prima Nota Cassa di {mese_nome} {prev_year} al commercialista!" if show_alert else None
    }


@router.get("/export-completo/{anno}/{mese}")
@handle_errors
async def export_dati_completi(anno: int, mese: int):
    """
    Export completo dati mensili per commercialista in formato ZIP.
    Include: fatture, corrispettivi, prima nota, IVA, dipendenti.
    """
    import zipfile
    import csv
    from io import BytesIO, StringIO
    from fastapi.responses import StreamingResponse
    
    db = Database.get_db()
    mese_str = f"{anno}-{mese:02d}"
    mese_nome = ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
                 "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"][mese]
    
    # Crea ZIP in memoria
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 1. FATTURE
        fatture = await db["invoices"].find({
            "invoice_date": {"$regex": f"^{mese_str}"}
        }, {"_id": 0}).to_list(10000)
        
        if fatture:
            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer, delimiter=';')
            writer.writerow(['Data', 'Numero', 'Fornitore', 'P.IVA', 'Imponibile', 'IVA', 'Totale', 'Pagamento'])
            
            for f in fatture:
                writer.writerow([
                    f.get('invoice_date', '')[:10],
                    f.get('invoice_number', ''),
                    f.get('supplier_name', ''),
                    f.get('supplier_vat', ''),
                    f.get('total_amount', 0) - f.get('total_tax', 0),
                    f.get('total_tax', 0),
                    f.get('total_amount', 0),
                    f.get('metodo_pagamento', 'N/D')
                ])
            
            zf.writestr(f'fatture_{mese_str}.csv', csv_buffer.getvalue())
        
        # 2. CORRISPETTIVI
        corrispettivi = await db["corrispettivi"].find({
            "$or": [
                {"data": {"$regex": f"^{mese_str}"}},
                {"data_trasmissione": {"$regex": f"^{mese_str}"}}
            ]
        }, {"_id": 0}).to_list(10000)
        
        if corrispettivi:
            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer, delimiter=';')
            writer.writerow(['Data', 'Totale', 'Contanti', 'Elettronico', 'IVA 10%', 'Matricola RT'])
            
            for c in corrispettivi:
                totale = c.get('totale', 0)
                iva = round(totale * 0.10 / 1.10, 2)
                writer.writerow([
                    c.get('data', c.get('data_trasmissione', ''))[:10],
                    totale,
                    c.get('pagato_contante', c.get('pagato_cassa', 0)),
                    c.get('pagato_elettronico', 0),
                    iva,
                    c.get('matricola_rt', '')
                ])
            
            zf.writestr(f'corrispettivi_{mese_str}.csv', csv_buffer.getvalue())
        
        # 3. PRIMA NOTA CASSA
        prima_nota_data = await get_prima_nota_cassa_mensile(anno, mese)
        
        if prima_nota_data.get('movimenti'):
            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer, delimiter=';')
            writer.writerow(['Data', 'Descrizione', 'Categoria', 'Entrata', 'Uscita', 'Tipo'])
            
            for m in prima_nota_data['movimenti']:
                importo = m.get('importo', 0)
                writer.writerow([
                    m.get('data', '')[:10],
                    m.get('descrizione', m.get('causale', '')),
                    m.get('categoria', ''),
                    importo if importo > 0 else '',
                    abs(importo) if importo < 0 else '',
                    m.get('tipo', '')
                ])
            
            zf.writestr(f'prima_nota_cassa_{mese_str}.csv', csv_buffer.getvalue())
        
        # 4. RIEPILOGO IVA
        totale_fatture = sum(f.get('total_amount', 0) for f in fatture)
        iva_credito = sum(f.get('total_tax', 0) for f in fatture)
        totale_corrispettivi = sum(c.get('totale', 0) for c in corrispettivi)
        iva_debito = round(totale_corrispettivi * 0.10 / 1.10, 2)
        saldo_iva = iva_debito - iva_credito
        
        riepilogo = f"""RIEPILOGO IVA - {mese_nome} {anno}
=====================================

ACQUISTI (Fatture):
  Numero Fatture: {len(fatture)}
  Totale Imponibile: € {totale_fatture - iva_credito:,.2f}
  IVA a Credito: € {iva_credito:,.2f}
  Totale Fatture: € {totale_fatture:,.2f}

VENDITE (Corrispettivi):
  Numero Corrispettivi: {len(corrispettivi)}
  Totale Incassato: € {totale_corrispettivi:,.2f}
  IVA a Debito (10%): € {iva_debito:,.2f}

SALDO IVA: € {saldo_iva:,.2f}
{'DA VERSARE' if saldo_iva > 0 else 'A CREDITO'}

=====================================
Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}
"""
        zf.writestr(f'riepilogo_iva_{mese_str}.txt', riepilogo)
        
        # 5. DIPENDENTI (se ci sono buste paga)
        buste_paga = await db["cedolini"].find({
            "mese": mese,
            "anno": anno
        }, {"_id": 0}).to_list(500)
        
        if buste_paga:
            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer, delimiter=';')
            writer.writerow(['Dipendente', 'Netto', 'Lordo', 'Contributi', 'Data Pagamento'])
            
            for b in buste_paga:
                writer.writerow([
                    b.get('dipendente_nome', ''),
                    b.get('netto', 0),
                    b.get('lordo', 0),
                    b.get('contributi', 0),
                    b.get('data_pagamento', '')
                ])
            
            zf.writestr(f'buste_paga_{mese_str}.csv', csv_buffer.getvalue())
    
    zip_buffer.seek(0)
    
    return StreamingResponse(
        zip_buffer,
        media_type='application/zip',
        headers={'Content-Disposition': f'attachment; filename=export_commercialista_{mese_str}.zip'}
    )


@router.get("/export-excel/{anno}/{mese}")
@handle_errors
async def export_excel_commercialista(anno: int, mese: int):
    """
    Export Excel mensile per commercialista.
    Include fogli separati per: fatture, corrispettivi, prima nota, IVA.
    Senza dettaglio deducibilità come richiesto.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    
    db = Database.get_db()
    mese_str = f"{anno}-{mese:02d}"
    mese_nome = ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
                 "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"][mese]
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    money_format = '#,##0.00 €'
    
    wb = Workbook()
    
    # === FOGLIO 1: FATTURE ACQUISTO ===
    ws_fatture = wb.active
    ws_fatture.title = "Fatture Acquisto"
    
    fatture = await db["invoices"].find({
        "invoice_date": {"$regex": f"^{mese_str}"}
    }, {"_id": 0}).sort("invoice_date", 1).to_list(10000)
    
    headers_fatture = ['Data', 'N. Fattura', 'Fornitore', 'P.IVA Fornitore', 'Categoria', 
                       'Imponibile', 'IVA', 'Totale', 'Pagamento', 'Conto']
    
    for col, header in enumerate(headers_fatture, 1):
        cell = ws_fatture.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    tot_imponibile = 0
    tot_iva = 0
    tot_fatture = 0
    
    for row, f in enumerate(fatture, 2):
        imponibile = float(f.get('total_amount', 0) or 0) - float(f.get('total_tax', 0) or 0)
        iva = float(f.get('total_tax', 0) or 0)
        totale = float(f.get('total_amount', 0) or 0)
        
        tot_imponibile += imponibile
        tot_iva += iva
        tot_fatture += totale
        
        values = [
            f.get('invoice_date', '')[:10],
            f.get('invoice_number', ''),
            f.get('supplier_name', ''),
            f.get('supplier_vat', ''),
            f.get('categoria_contabile', '').replace('_', ' ').title(),
            imponibile,
            iva,
            totale,
            f.get('payment_method', f.get('metodo_pagamento', '')),
            f.get('conto_costo_codice', '')
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws_fatture.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [6, 7, 8]:
                cell.number_format = money_format
    
    # Riga totali
    row_tot = len(fatture) + 2
    ws_fatture.cell(row=row_tot, column=5, value="TOTALI:").font = Font(bold=True)
    ws_fatture.cell(row=row_tot, column=6, value=tot_imponibile).number_format = money_format
    ws_fatture.cell(row=row_tot, column=7, value=tot_iva).number_format = money_format
    ws_fatture.cell(row=row_tot, column=8, value=tot_fatture).number_format = money_format
    for col in [6, 7, 8]:
        ws_fatture.cell(row=row_tot, column=col).font = Font(bold=True)
    
    # Larghezza colonne
    for col in range(1, 11):
        ws_fatture.column_dimensions[get_column_letter(col)].width = 15 if col not in [3] else 30
    
    # === FOGLIO 2: CORRISPETTIVI ===
    ws_corr = wb.create_sheet("Corrispettivi")
    
    corrispettivi = await db["corrispettivi"].find({
        "data": {"$regex": f"^{mese_str}"}
    }, {"_id": 0}).sort("data", 1).to_list(10000)
    
    headers_corr = ['Data', 'Totale', 'Contante', 'Elettronico', 'Chiusura N.', 'Note']
    
    for col, header in enumerate(headers_corr, 1):
        cell = ws_corr.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    tot_corr = 0
    tot_contante = 0
    tot_elettr = 0
    
    for row, c in enumerate(corrispettivi, 2):
        totale = float(c.get('totale', 0) or 0)
        contante = float(c.get('pagato_contante', c.get('pagato_cassa', 0)) or 0)
        elettr = float(c.get('pagato_elettronico', 0) or 0)
        
        tot_corr += totale
        tot_contante += contante
        tot_elettr += elettr
        
        values = [
            c.get('data', '')[:10],
            totale,
            contante,
            elettr,
            c.get('numero_chiusura', ''),
            c.get('note', '')
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws_corr.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [2, 3, 4]:
                cell.number_format = money_format
    
    # Totali corrispettivi
    row_tot = len(corrispettivi) + 2
    ws_corr.cell(row=row_tot, column=1, value="TOTALI:").font = Font(bold=True)
    ws_corr.cell(row=row_tot, column=2, value=tot_corr).number_format = money_format
    ws_corr.cell(row=row_tot, column=3, value=tot_contante).number_format = money_format
    ws_corr.cell(row=row_tot, column=4, value=tot_elettr).number_format = money_format
    
    for col in range(1, 7):
        ws_corr.column_dimensions[get_column_letter(col)].width = 15
    
    # === FOGLIO 3: PRIMA NOTA CASSA ===
    ws_pn = wb.create_sheet("Prima Nota Cassa")
    
    # Ordinamento: prima per data, poi per categoria (Corrispettivi prima di POS)
    prima_nota = await db["prima_nota_cassa"].find({
        "data": {"$regex": f"^{mese_str}"}
    }, {"_id": 0}).sort([("data", 1), ("categoria", 1)]).to_list(10000)
    
    headers_pn = ['Data', 'Descrizione', 'Categoria', 'Tipo', 'Importo']
    
    for col, header in enumerate(headers_pn, 1):
        cell = ws_pn.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    tot_entrate = 0
    tot_uscite = 0
    
    for row, pn in enumerate(prima_nota, 2):
        importo = float(pn.get('importo', 0) or 0)
        tipo = pn.get('tipo', 'uscita')
        
        if tipo == 'entrata':
            tot_entrate += importo
        else:
            tot_uscite += importo
        
        values = [
            pn.get('data', '')[:10],
            pn.get('descrizione', ''),
            pn.get('categoria', '').replace('_', ' ').title(),
            tipo.upper(),
            importo
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws_pn.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 5:
                cell.number_format = money_format
    
    # Totali prima nota
    row_tot = len(prima_nota) + 2
    ws_pn.cell(row=row_tot, column=3, value="TOTALE ENTRATE:").font = Font(bold=True)
    ws_pn.cell(row=row_tot, column=5, value=tot_entrate).number_format = money_format
    ws_pn.cell(row=row_tot + 1, column=3, value="TOTALE USCITE:").font = Font(bold=True)
    ws_pn.cell(row=row_tot + 1, column=5, value=tot_uscite).number_format = money_format
    ws_pn.cell(row=row_tot + 2, column=3, value="SALDO:").font = Font(bold=True, color="0000FF")
    ws_pn.cell(row=row_tot + 2, column=5, value=tot_entrate - tot_uscite).number_format = money_format
    
    ws_pn.column_dimensions['B'].width = 40
    for col in [1, 3, 4, 5]:
        ws_pn.column_dimensions[get_column_letter(col)].width = 15
    
    # === FOGLIO 4: RIEPILOGO IVA ===
    ws_iva = wb.create_sheet("Riepilogo IVA")
    
    headers_iva = ['Voce', 'Importo']
    for col, header in enumerate(headers_iva, 1):
        cell = ws_iva.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # IVA vendite (dai corrispettivi - assumiamo 10%)
    iva_vendite = tot_corr * 0.10 / 1.10
    
    iva_data = [
        ('IVA a debito (vendite)', iva_vendite),
        ('IVA a credito (acquisti)', tot_iva),
        ('', ''),
        ('SALDO IVA', iva_vendite - tot_iva)
    ]
    
    for row, (voce, importo) in enumerate(iva_data, 2):
        ws_iva.cell(row=row, column=1, value=voce).border = border
        if importo != '':
            cell = ws_iva.cell(row=row, column=2, value=importo)
            cell.number_format = money_format
            cell.border = border
        if voce == 'SALDO IVA':
            ws_iva.cell(row=row, column=1).font = Font(bold=True)
            ws_iva.cell(row=row, column=2).font = Font(bold=True)
            if importo > 0:
                ws_iva.cell(row=row, column=2).font = Font(bold=True, color="FF0000")
    
    ws_iva.column_dimensions['A'].width = 30
    ws_iva.column_dimensions['B'].width = 20
    
    # === FOGLIO 5: RIEPILOGO GENERALE ===
    ws_riep = wb.create_sheet("Riepilogo")
    
    ws_riep.cell(row=1, column=1, value=f"RIEPILOGO {mese_nome.upper()} {anno}").font = Font(bold=True, size=14)
    ws_riep.merge_cells('A1:B1')
    
    riepilogo = [
        ('', ''),
        ('FATTURE ACQUISTO', ''),
        ('  Numero fatture', len(fatture)),
        ('  Totale imponibile', tot_imponibile),
        ('  Totale IVA', tot_iva),
        ('  Totale fatture', tot_fatture),
        ('', ''),
        ('CORRISPETTIVI', ''),
        ('  Giorni registrati', len(corrispettivi)),
        ('  Totale incassato', tot_corr),
        ('  di cui contante', tot_contante),
        ('  di cui elettronico', tot_elettr),
        ('', ''),
        ('PRIMA NOTA CASSA', ''),
        ('  Totale entrate', tot_entrate),
        ('  Totale uscite', tot_uscite),
        ('  Saldo cassa', tot_entrate - tot_uscite),
        ('', ''),
        ('IVA', ''),
        ('  IVA a debito', iva_vendite),
        ('  IVA a credito', tot_iva),
        ('  Saldo IVA', iva_vendite - tot_iva)
    ]
    
    for row, (voce, valore) in enumerate(riepilogo, 3):
        cell_voce = ws_riep.cell(row=row, column=1, value=voce)
        if valore != '' and not voce.startswith('  '):
            cell_voce.font = Font(bold=True)
        if valore != '':
            cell_val = ws_riep.cell(row=row, column=2, value=valore)
            if isinstance(valore, (int, float)):
                cell_val.number_format = money_format if isinstance(valore, float) else '0'
    
    ws_riep.column_dimensions['A'].width = 25
    ws_riep.column_dimensions['B'].width = 18
    
    # Genera file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"contabilita_{mese_nome.lower()}_{anno}.xlsx"
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )



# ============================================
# EXPORT SCHEDULATO AUTOMATICO
# ============================================

@router.post("/schedula-export")
@handle_errors
async def schedula_export_mensile(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """
    Schedula o esegue immediatamente l'invio del report mensile al commercialista.
    
    Body:
    - anno: int
    - mese: int (1-12)
    - email: str (opzionale, usa default se non specificato)
    - immediato: bool (se True, invia subito invece di schedulare)
    """
    db = Database.get_db()
    
    anno = data.get("anno", datetime.now().year)
    mese = data.get("mese", datetime.now().month)
    email = data.get("email")
    immediato = data.get("immediato", True)
    
    # Recupera config
    config = await db["commercialista_config"].find_one({}, {"_id": 0})
    if not email:
        email = config.get("email") if config else DEFAULT_COMMERCIALISTA_EMAIL
    
    MESI_NOMI = ['Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno', 
                 'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre']
    mese_nome = MESI_NOMI[mese - 1]
    
    if immediato:
        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from datetime import timezone
            
            # Genera il report Excel
            _, ultimo_giorno = monthrange(anno, mese)
            data_inizio = f"{anno}-{mese:02d}-01"
            data_fine = f"{anno}-{mese:02d}-{ultimo_giorno:02d}"
            
            # Raccogli dati
            fatture = await db["invoices"].find({
                "data_ricezione": {"$gte": data_inizio, "$lte": data_fine}
            }, {"_id": 0}).to_list(1000)
            
            corrispettivi = await db["corrispettivi"].find({
                "data": {"$gte": data_inizio, "$lte": data_fine}
            }, {"_id": 0}).to_list(100)
            
            prima_nota = await db["prima_nota_cassa"].find({
                "data": {"$gte": data_inizio, "$lte": data_fine}
            }, {"_id": 0}).to_list(1000)
            
            # Crea Excel semplificato
            wb = Workbook()
            ws = wb.active
            ws.title = "Riepilogo"
            
            # Titolo
            ws['A1'] = f"REPORT MENSILE - {mese_nome} {anno}"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')
            
            # Statistiche
            tot_fatture = sum(float(f.get("totale", 0) or 0) for f in fatture)
            tot_corr = sum(float(c.get("totale", 0) or 0) for c in corrispettivi)
            entrate = sum(float(m.get("importo", 0) or 0) for m in prima_nota if m.get("tipo") == "entrata")
            uscite = sum(abs(float(m.get("importo", 0) or 0)) for m in prima_nota if m.get("tipo") == "uscita")
            
            stats = [
                ("", ""),
                ("FATTURE ACQUISTO", ""),
                ("  Numero fatture", len(fatture)),
                ("  Totale fatture", f"€ {tot_fatture:,.2f}"),
                ("", ""),
                ("CORRISPETTIVI", ""),
                ("  Giorni registrati", len(corrispettivi)),
                ("  Totale incassato", f"€ {tot_corr:,.2f}"),
                ("", ""),
                ("PRIMA NOTA", ""),
                ("  Entrate", f"€ {entrate:,.2f}"),
                ("  Uscite", f"€ {uscite:,.2f}"),
                ("  Saldo", f"€ {entrate - uscite:,.2f}"),
            ]
            
            for idx, (label, value) in enumerate(stats, 3):
                ws.cell(row=idx, column=1, value=label)
                ws.cell(row=idx, column=2, value=value)
            
            # Genera bytes
            output = io.BytesIO()
            wb.save(output)
            excel_bytes = output.getvalue()
            
            # Invia email
            subject = f"📊 Report Contabile {mese_nome} {anno} - Azienda in Cloud"
            html_body = f"""
            <html>
            <body style="font-family: Arial, sans-serif; color: #333;">
                <h2>📊 Report Mensile - {mese_nome} {anno}</h2>
                <p>In allegato il report contabile mensile con:</p>
                <ul>
                    <li><strong>Fatture acquisto:</strong> {len(fatture)} documenti (€ {tot_fatture:,.2f})</li>
                    <li><strong>Corrispettivi:</strong> {len(corrispettivi)} giorni (€ {tot_corr:,.2f})</li>
                    <li><strong>Prima Nota:</strong> Entrate € {entrate:,.2f} / Uscite € {uscite:,.2f}</li>
                </ul>
                <p style="color: #666; font-size: 12px;">
                    Report generato automaticamente da Azienda in Cloud ERP<br>
                    Data invio: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}
                </p>
            </body>
            </html>
            """
            
            filename = f"report_{mese_nome.lower()}_{anno}.xlsx"
            
            send_email_with_attachment(
                to_email=email,
                subject=subject,
                html_body=html_body,
                attachment_data=excel_bytes,
                attachment_name=filename
            )
            
            # Salva log
            log_export_doc = {
                "tipo": "report_mensile",
                "anno": anno,
                "mese": mese,
                "email": email,
                "inviato_at": datetime.now(timezone.utc).isoformat(),
                "statistiche": {
                    "fatture": len(fatture),
                    "corrispettivi": len(corrispettivi),
                    "prima_nota": len(prima_nota)
                }
            }
            await db["export_log"].insert_one(log_export_doc.copy())
            
            return {
                "success": True,
                "message": f"Report {mese_nome} {anno} inviato a {email}",
                "email": email,
                "statistiche": {
                    "fatture": len(fatture),
                    "tot_fatture": tot_fatture,
                    "corrispettivi": len(corrispettivi),
                    "entrate": entrate,
                    "uscite": uscite
                }
            }
            
        except Exception as e:
            logger.error(f"Errore invio report: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    else:
        # Schedula per il futuro (salva in DB)
        scheduled_doc = {
            "tipo": "report_mensile",
            "anno": anno,
            "mese": mese,
            "email": email,
            "scheduled_at": datetime.now(timezone.utc).isoformat(),
            "status": "pending"
        }
        await db["scheduled_exports"].insert_one(scheduled_doc.copy())
        
        return {
            "success": True,
            "message": f"Export schedulato per {mese_nome} {anno}",
            "scheduled": True
        }


@router.get("/export-log")
@handle_errors
async def get_export_log(limit: int = 20) -> Dict[str, Any]:
    """Recupera lo storico degli export inviati."""
    db = Database.get_db()
    
    logs = await db["export_log"].find(
        {},
        {"_id": 0}
    ).sort("inviato_at", -1).limit(limit).to_list(limit)
    
    return {
        "logs": logs,
        "total": await db["export_log"].count_documents({})
    }
