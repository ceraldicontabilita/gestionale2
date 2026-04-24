"""
Cash router.
API endpoints for cash register management and daily closures.
"""
from fastapi import APIRouter, Depends, Query, Path, status
from fastapi.responses import StreamingResponse
import io
from typing import List, Dict, Any, Optional
from datetime import date
import logging

from app.database import Database, Collections
from app.repositories.cash_repository import (
    CashMovementRepository,
)
from app.services.cash_service import CashService
from app.models.cash import (
    CashMovementCreate,
    CashMovementUpdate,
    CorrissettivoCreate,
    CashStats
)
from app.utils.dependencies import get_current_user, pagination_params

logger = logging.getLogger(__name__)

router = APIRouter()


# Dependency to get cash service
async def get_cash_service() -> CashService:
    """Get cash service with injected dependencies."""
    db = Database.get_db()
    movement_repo = CashMovementRepository(db[Collections.CASH_MOVEMENTS])
    
    return CashService(movement_repo, None)


@router.get(
    "/movements",
    response_model=List[Dict[str, Any]],
    summary="List cash movements",
    description="Get list of cash register movements with filters"
)
async def list_movements(
    current_user: Dict[str, Any] = Depends(get_current_user),
    pagination: Dict[str, Any] = Depends(pagination_params),
    start_date: Optional[date] = Query(None, description="Start date filter"),
    end_date: Optional[date] = Query(None, description="End date filter"),
    type: Optional[str] = Query(None, description="Type: entrata or uscita"),
    category: Optional[str] = Query(None, description="Category filter"),
    service: CashService = Depends(get_cash_service)
) -> List[Dict[str, Any]]:
    """
    List cash movements with filters.
    
    **Query Parameters:**
    - **skip**: Pagination offset
    - **limit**: Maximum results
    - **start_date**: Filter from this date (YYYY-MM-DD)
    - **end_date**: Filter to this date (YYYY-MM-DD)
    - **type**: Filter by type (entrata/uscita)
    - **category**: Filter by category
    
    **Returns:**
    - List of cash movements (newest first)
    """
    user_id = current_user["user_id"]
    
    return await service.list_movements(
        user_id=user_id,
        start_date=start_date,
        end_date=end_date,
        movement_type=type,
        category=category,
        skip=pagination["skip"],
        limit=pagination["limit"]
    )


@router.post(
    "/movements",
    response_model=Dict[str, str],
    status_code=status.HTTP_201_CREATED,
    summary="Create cash movement",
    description="Create a new cash register movement"
)
async def create_movement(
    movement_data: CashMovementCreate,
    current_user: Dict[str, Any] = Depends(get_current_user),
    service: CashService = Depends(get_cash_service)
) -> Dict[str, str]:
    """
    Create cash movement.
    
    **Request Body:**
    - **date**: Movement date (YYYY-MM-DD)
    - **time**: Movement time (HH:MM)
    - **type**: Movement type (entrata/uscita)
    - **amount**: Amount in euros
    - **description**: Movement description
    - **category**: Optional category
    - **payment_method**: Optional payment method (contanti/carta/assegno)
    - **reference_number**: Optional receipt/invoice number
    - **notes**: Optional notes
    
    **Returns:**
    - Movement ID and success message
    """
    user_id = current_user["user_id"]
    
    movement_id = await service.create_movement(
        movement_data=movement_data,
        user_id=user_id
    )
    
    return {
        "message": "Cash movement created successfully",
        "movement_id": movement_id
    }


@router.put(
    "/movements/{movement_id}",
    status_code=status.HTTP_200_OK,
    summary="Update cash movement",
    description="Update an existing cash movement"
)
async def update_movement(
    movement_id: str = Path(..., description="Movement ID"),
    update_data: CashMovementUpdate = ...,
    current_user: Dict[str, Any] = Depends(get_current_user),
    service: CashService = Depends(get_cash_service)
) -> Dict[str, str]:
    """
    Update cash movement.
    
    Only provided fields will be updated.
    
    **Path Parameters:**
    - **movement_id**: Movement ID
    
    **Request Body:** (all optional)
    - **amount**: Amount
    - **description**: Description
    - **category**: Category
    - **payment_method**: Payment method
    - **reference_number**: Reference number
    - **notes**: Notes
    
    **Raises:**
    - 404: If movement not found
    """
    await service.update_movement(
        movement_id=movement_id,
        update_data=update_data
    )
    
    return {"message": "Cash movement updated successfully"}


@router.delete(
    "/movements/{movement_id}",
    status_code=status.HTTP_200_OK,
    summary="Delete cash movement",
    description="Delete a cash movement"
)
async def delete_movement(
    movement_id: str = Path(..., description="Movement ID"),
    current_user: Dict[str, Any] = Depends(get_current_user),
    service: CashService = Depends(get_cash_service)
) -> Dict[str, str]:
    """
    Delete a cash movement.
    
    **Path Parameters:**
    - **movement_id**: Movement ID
    
    **Raises:**
    - 404: If movement not found
    """
    await service.delete_movement(movement_id)
    
    return {"message": "Cash movement deleted successfully"}


@router.get(
    "/stats",
    response_model=CashStats,
    summary="Get cash statistics",
    description="Get cash register statistics for date range"
)
async def get_cash_stats(
    start_date: date = Query(..., description="Start date"),
    end_date: date = Query(..., description="End date"),
    current_user: Dict[str, Any] = Depends(get_current_user),
    service: CashService = Depends(get_cash_service)
) -> Dict[str, Any]:
    """
    Get cash register statistics.
    
    **Query Parameters:**
    - **start_date**: Start date (YYYY-MM-DD)
    - **end_date**: End date (YYYY-MM-DD)
    
    **Returns:**
    - total_in: Total inflows
    - total_out: Total outflows
    - current_balance: Balance (in - out)
    - movements_count: Number of movements
    - by_category: Breakdown by category
    - by_payment_method: Breakdown by payment method
    """
    user_id = current_user["user_id"]
    
    return await service.get_cash_stats(
        user_id=user_id,
        start_date=start_date,
        end_date=end_date
    )


@router.post(
    "/corrispettivi",
    response_model=Dict[str, str],
    status_code=status.HTTP_201_CREATED,
    summary="Create corrispettivo",
    description="Create daily cash register closure (corrispettivo)"
)
async def create_corrispettivo(
    corrispettivo_data: CorrissettivoCreate,
    current_user: Dict[str, Any] = Depends(get_current_user),
    service: CashService = Depends(get_cash_service)
) -> Dict[str, str]:
    """
    Create corrispettivo (daily closure).
    
    Required for Italian tax compliance.
    
    **Request Body:**
    - **date**: Closure date (YYYY-MM-DD)
    - **total_cash**: Total cash payments
    - **total_card**: Total card payments
    - **total_other**: Other payment methods
    - **total_receipts**: Number of receipts issued
    - **opening_balance**: Optional opening cash balance
    - **closing_balance**: Optional closing cash balance (counted)
    - **notes**: Optional notes
    
    **Returns:**
    - Corrispettivo ID and success message
    
    **Notes:**
    - If opening_balance provided, system calculates expected_balance
    - If closing_balance provided, system calculates difference
    
    **Raises:**
    - 400: If corrispettivo already exists for date
    """
    user_id = current_user["user_id"]
    
    corrispettivo_id = await service.create_corrispettivo(
        corrispettivo_data=corrispettivo_data,
        user_id=user_id
    )
    
    return {
        "message": "Corrispettivo created successfully",
        "corrispettivo_id": corrispettivo_id
    }


@router.get(
    "/corrispettivi/{target_date}",
    response_model=Optional[Dict[str, Any]],
    summary="Get corrispettivo",
    description="Get corrispettivo for specific date"
)
async def get_corrispettivo(
    target_date: date = Path(..., description="Date (YYYY-MM-DD)"),
    current_user: Dict[str, Any] = Depends(get_current_user),
    service: CashService = Depends(get_cash_service)
) -> Optional[Dict[str, Any]]:
    """
    Get corrispettivo for specific date.
    
    **Path Parameters:**
    - **target_date**: Date (YYYY-MM-DD)
    
    **Returns:**
    - Corrispettivo document or null if not found
    """
    user_id = current_user["user_id"]
    
    return await service.get_corrispettivo(
        corrispettivo_date=target_date,
        user_id=user_id
    )


@router.get(
    "/export/excel",
    summary="Export cash movements to Excel",
    description="Export cash movements to Excel file"
)
async def export_cash_excel(
    start_date: date = Query(..., description="Start date"),
    end_date: date = Query(..., description="End date"),
    current_user: Dict[str, Any] = Depends(get_current_user),
    service: CashService = Depends(get_cash_service)
):
    """
    Export movimenti cassa in Excel (openpyxl).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=501,
            content={"error": "openpyxl non installato", "detail": "pip install openpyxl"}
        )

    db = Database.get_db()
    query = {
        "data": {
            "$gte": start_date.isoformat(),
            "$lte": end_date.isoformat()
        }
    }
    movimenti = await db[Collections.CASH_MOVEMENTS].find(
        query, {"_id": 0}
    ).sort("data", 1).to_list(10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimenti Cassa"

    # Header
    headers = ["Data", "Tipo", "Descrizione", "Causale", "Importo (€)", "Saldo"]
    header_fill = PatternFill("solid", fgColor="0F2744")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Dati
    saldo = 0.0
    for row_idx, mov in enumerate(movimenti, 2):
        importo = float(mov.get("importo") or 0)
        tipo = mov.get("tipo", "")
        if tipo in ("entrata", "incasso", "+"):
            saldo += importo
        else:
            saldo -= importo

        ws.cell(row=row_idx, column=1, value=mov.get("data", ""))
        ws.cell(row=row_idx, column=2, value=tipo)
        ws.cell(row=row_idx, column=3, value=mov.get("descrizione") or mov.get("note", ""))
        ws.cell(row=row_idx, column=4, value=mov.get("causale", ""))
        ws.cell(row=row_idx, column=5, value=importo)
        ws.cell(row=row_idx, column=6, value=round(saldo, 2))

    # Larghezze colonne
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20

    # Salva in buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"cassa_{start_date}_{end_date}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
