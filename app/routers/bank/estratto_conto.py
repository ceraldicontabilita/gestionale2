"""
Gestione Estratto Conto
Salva e visualizza tutti i movimenti bancari importati con campi strutturati.
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from typing import Dict, Any, List, Optional
from datetime import datetime, date
import logging
import io
import re
import csv

from app.database import Database
from app.utils.error_handler import handle_errors

logger = logging.getLogger(__name__)
router = APIRouter()


def estrai_numero_fattura(descrizione: str) -> Optional[str]:
    """Estrae il numero/i fattura dalla descrizione dopo NOTPROVIDE."""
    if not descrizione:
        return None
    
    # Pattern: dopo "NOTPROVIDE - " c'è il riferimento fatture
    match = re.search(r'NOTPROVIDE\s*-?\s*(.+)$', descrizione, re.IGNORECASE)
    if match:
        riferimento = match.group(1).strip()
        # Pulisci e restituisci
        # Rimuovi prefissi comuni
        riferimento = re.sub(r'^(saldo|pagamento)\s+(fattur[ae]|ft)\s*', '', riferimento, flags=re.IGNORECASE)
        return riferimento.strip()[:200] if riferimento.strip() else None
    
    # Pattern alternativo: "fattura/e" seguito da numeri
    match = re.search(r'fattur[ae]\s+(.+?)(?:\s*$|\s+-)', descrizione, re.IGNORECASE)
    if match:
        return match.group(1).strip()[:200]
    
    return None


def estrai_fornitore_pulito(descrizione: str) -> Optional[str]:
    """Estrae il nome fornitore dalla descrizione, pulendolo."""
    if not descrizione:
        return None
    
    desc_upper = descrizione.upper()
    
    if "FAVORE" in desc_upper:
        idx = desc_upper.find("FAVORE")
        after = descrizione[idx + 7:].strip()
        
        # Prendi fino a "NOTPROVIDE" o " - " o fine
        for sep in ["NOTPROVIDE", " - ADD.", " - "]:
            if sep.upper() in after.upper():
                idx_sep = after.upper().find(sep.upper())
                after = after[:idx_sep].strip()
                break
        
        # Rimuovi forme societarie alla fine per pulizia display
        # Ma mantienile se sono parte del nome
        nome = after.strip()
        
        return nome if nome else None
    
    return None


@router.post("/import")
@handle_errors
async def import_estratto_conto(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa estratto conto bancario e salva tutti i movimenti con campi strutturati.
    
    Formato CSV atteso (delimitatore ';'):
    - Ragione Sociale: nome azienda
    - Data contabile: data in formato DD/MM/YYYY
    - Data valuta: data valuta in formato DD/MM/YYYY
    - Banca: nome banca e codice
    - Rapporto: numero rapporto/conto
    - Importo: importo con virgola decimale (es: 254,5)
    - Divisa: valuta (EUR)
    - Descrizione: descrizione del movimento
    - Categoria/sottocategoria: categoria del movimento
    - Hashtag: tag opzionale
    
    Evita duplicati controllando data + importo + descrizione.
    """
    db = Database.get_db()
    
    filename = file.filename.lower()
    contents = await file.read()
    
    movimenti = []
    
    if filename.endswith('.csv'):
        # Prova diversi encoding
        text = None
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                text = contents.decode(encoding)
                break
            except (UnicodeDecodeError, Exception):
                continue
        
        if not text:
            raise HTTPException(status_code=400, detail="Impossibile decodificare il file CSV")
        
        reader = csv.DictReader(io.StringIO(text), delimiter=';')
        
        for row in reader:
            # Estrai dati con supporto per varianti di nomi colonna
            # Supporta sia formato con virgolette che senza
            
            # Ragione sociale
            ragione_sociale = row.get('Ragione Sociale', '') or row.get('"Ragione Sociale"', '')
            
            # Data contabile
            data_contabile = (
                row.get('Data contabile', '') or 
                row.get('"Data contabile"', '') or
                row.get('Data', '')
            ).strip().strip('"')
            
            # Data valuta
            data_valuta = (
                row.get('Data valuta', '') or 
                row.get('"Data valuta"', '') or
                row.get('Data valut.', '')
            ).strip().strip('"')
            
            # Banca
            banca = (row.get('Banca', '') or row.get('"Banca"', '')).strip().strip('"')
            
            # Rapporto (numero conto)
            rapporto = (row.get('Rapporto', '') or row.get('"Rapporto"', '')).strip().strip('"')
            
            # Importo - può essere con virgola come separatore decimale
            importo_str = (
                row.get('Importo', '0') or 
                row.get('"Importo"', '0')
            ).strip().strip('"')
            
            # Parse importo: rimuovi punti migliaia, sostituisci virgola con punto
            importo_str = importo_str.replace('.', '').replace(',', '.')
            try:
                importo = float(importo_str)
            except (ValueError, TypeError):
                continue
            
            # Divisa
            divisa = (row.get('Divisa', 'EUR') or row.get('"Divisa"', 'EUR')).strip().strip('"')
            
            # Descrizione
            descrizione = (
                row.get('Descrizione', '') or 
                row.get('"Descrizione"', '') or
                row.get('Descrizion', '')
            ).strip().strip('"')
            
            # Categoria
            categoria = (
                row.get('Categoria/sottocategoria', '') or 
                row.get('"Categoria/sottocategoria"', '') or
                row.get('Categoria', '') or
                row.get('"Categoria"', '')
            ).strip().strip('"')
            
            # Hashtag
            hashtag = (row.get('Hashtag', '') or row.get('"Hashtag"', '')).strip().strip('"')
            
            # Parse data contabile (DD/MM/YYYY)
            try:
                if '/' in data_contabile:
                    parts = data_contabile.split('/')
                    data_obj = date(int(parts[2]), int(parts[1]), int(parts[0]))
                else:
                    continue
            except (ValueError, TypeError, IndexError):
                continue
            
            # Parse data valuta
            data_pagamento = None
            try:
                if '/' in data_valuta:
                    parts = data_valuta.split('/')
                    data_pagamento = date(int(parts[2]), int(parts[1]), int(parts[0]))
            except (ValueError, TypeError, IndexError):
                pass
            
            # Estrai fornitore/beneficiario dalla descrizione
            fornitore = estrai_fornitore_pulito(descrizione)
            
            # Estrai numero fattura dalla descrizione
            numero_fattura = estrai_numero_fattura(descrizione)
            
            movimenti.append({
                "data": data_obj,
                "ragione_sociale": ragione_sociale.strip().strip('"') if ragione_sociale else None,
                "fornitore": fornitore,
                "importo": importo,
                "numero_fattura": numero_fattura,
                "data_pagamento": data_pagamento,
                "categoria": categoria,
                "descrizione_originale": descrizione,
                "banca": banca,
                "rapporto": rapporto,
                "divisa": divisa,
                "hashtag": hashtag,
                "tipo": "uscita" if importo < 0 else "entrata"
            })
    
    elif filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            sheet = wb.active
            
            # Mappa header originali a chiavi normalizzate
            headers_raw = [str(cell.value or '') for cell in sheet[1]]
            headers = [h.lower().strip() for h in headers_raw]
            
            for row_num in range(2, sheet.max_row + 1):
                row_data = {headers[i]: sheet.cell(row=row_num, column=i+1).value 
                           for i in range(len(headers))}
                
                # Trova colonne con supporto per varianti
                data_contabile = None
                importo = None
                descrizione = ""
                categoria = ""
                data_valuta = None
                ragione_sociale = ""
                banca = ""
                rapporto = ""
                divisa = "EUR"
                hashtag = ""
                
                for h, v in row_data.items():
                    if not v:
                        continue
                    h_lower = h.lower()
                    
                    # Ragione Sociale
                    if 'ragione sociale' in h_lower:
                        ragione_sociale = str(v).strip()
                    
                    # Data contabile
                    elif 'data contabile' in h_lower or (h_lower == 'data' and not data_contabile):
                        if isinstance(v, (datetime, date)):
                            data_contabile = v if isinstance(v, date) else v.date()
                        elif '/' in str(v):
                            parts = str(v).split('/')
                            try:
                                data_contabile = date(int(parts[2]), int(parts[1]), int(parts[0]))
                            except (ValueError, TypeError, IndexError):
                                pass
                    
                    # Data valuta
                    elif 'data valuta' in h_lower or 'data valut' in h_lower:
                        if isinstance(v, (datetime, date)):
                            data_valuta = v if isinstance(v, date) else v.date()
                        elif '/' in str(v):
                            parts = str(v).split('/')
                            try:
                                data_valuta = date(int(parts[2]), int(parts[1]), int(parts[0]))
                            except (ValueError, TypeError, IndexError):
                                pass
                    
                    # Banca
                    elif h_lower == 'banca':
                        banca = str(v).strip()
                    
                    # Rapporto
                    elif h_lower == 'rapporto':
                        rapporto = str(v).strip()
                    
                    # Importo
                    elif 'importo' in h_lower:
                        if isinstance(v, (int, float)):
                            importo = float(v)
                        else:
                            try:
                                importo = float(str(v).replace('.', '').replace(',', '.'))
                            except (ValueError, TypeError):
                                pass
                    
                    # Divisa
                    elif h_lower == 'divisa':
                        divisa = str(v).strip()
                    
                    # Descrizione
                    elif 'descri' in h_lower:
                        descrizione = str(v).strip()
                    
                    # Categoria
                    elif 'categoria' in h_lower:
                        categoria = str(v).strip()
                    
                    # Hashtag
                    elif h_lower == 'hashtag':
                        hashtag = str(v).strip()
                
                if data_contabile and importo is not None:
                    movimenti.append({
                        "data": data_contabile,
                        "ragione_sociale": ragione_sociale if ragione_sociale else None,
                        "fornitore": estrai_fornitore_pulito(descrizione),
                        "importo": importo,
                        "numero_fattura": estrai_numero_fattura(descrizione),
                        "data_pagamento": data_valuta,
                        "categoria": categoria,
                        "descrizione_originale": descrizione,
                        "banca": banca,
                        "rapporto": rapporto,
                        "divisa": divisa,
                        "hashtag": hashtag,
                        "tipo": "uscita" if importo < 0 else "entrata"
                    })
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Errore parsing Excel: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Formato non supportato. Usa CSV o Excel.")
    
    # Salva nel database, evitando duplicati
    inserted = 0
    duplicates = 0
    
    for mov in movimenti:
        # Crea ID univoco basato su data + importo + primi 50 char descrizione
        desc_hash = (mov["descrizione_originale"] or "")[:50]
        mov_id = f"EC-{mov['data'].isoformat()}-{mov['importo']:.2f}-{hash(desc_hash) % 100000:05d}"
        
        # Controlla duplicati
        existing = await db["estratto_conto_movimenti"].find_one({
            "data": mov["data"].isoformat(),
            "importo": mov["importo"],
            "descrizione_hash": desc_hash
        })
        
        if existing:
            duplicates += 1
            continue
        
        # Salva
        record = {
            "id": mov_id,
            "data": mov["data"].isoformat(),
            "ragione_sociale": mov.get("ragione_sociale"),
            "fornitore": mov["fornitore"],
            "importo": mov["importo"],
            "numero_fattura": mov["numero_fattura"],
            "data_pagamento": mov["data_pagamento"].isoformat() if mov["data_pagamento"] else None,
            "categoria": mov["categoria"],
            "descrizione_originale": mov["descrizione_originale"],
            "banca": mov.get("banca"),
            "rapporto": mov.get("rapporto"),
            "divisa": mov.get("divisa", "EUR"),
            "hashtag": mov.get("hashtag"),
            "tipo": mov["tipo"],
            "descrizione_hash": desc_hash,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db["estratto_conto_movimenti"].insert_one(record.copy())
        inserted += 1
    
    # ===== RICONCILIAZIONE AUTOMATICA =====
    # Dopo l'import, avvia la riconciliazione automatica
    riconciliazione_results = None
    try:
        from app.routers.accounting.riconciliazione_automatica import riconcilia_estratto_conto
        riconciliazione_results = await riconcilia_estratto_conto()
    except Exception as e:
        logger.error(f"Errore riconciliazione automatica: {e}")
        riconciliazione_results = {"error": str(e)}
    
    # ===== RICONCILIAZIONE AUTOMATICA PAGHE (Stipendi + F24) =====
    riconciliazione_paghe = None
    try:
        from app.services.paghe_riconciliazione import esegui_riconciliazione_paghe_completa
        riconciliazione_paghe = await esegui_riconciliazione_paghe_completa(db)
    except Exception as e:
        logger.error(f"Errore riconciliazione paghe: {e}")
        riconciliazione_paghe = {"error": str(e)}
    
    return {
        "message": "Importazione estratto conto completata",
        "movimenti_trovati": len(movimenti),
        "movimenti_importati": inserted,
        "inseriti": inserted,
        "duplicati_saltati": duplicates,
        "riconciliazione_automatica": riconciliazione_results,
        "riconciliazione_paghe": riconciliazione_paghe
    }


@router.get("/movimenti")
@handle_errors
async def get_movimenti(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    categoria: Optional[str] = Query(None),
    fornitore: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None),  # "entrata" | "uscita"
    limit: int = Query(500, le=10000),
    offset: int = Query(0)
) -> Dict[str, Any]:
    """
    Recupera i movimenti dell'estratto conto con filtri.
    Ordinati per data decrescente.
    """
    db = Database.get_db()
    
    query = {}
    
    if anno:
        # Filtra per anno nella data (formato ISO: YYYY-MM-DD)
        query["data"] = {"$regex": f"^{anno}"}
        if mese:
            query["data"] = {"$regex": f"^{anno}-{mese:02d}"}
    
    if categoria:
        query["categoria"] = {"$regex": categoria, "$options": "i"}
    
    if fornitore:
        query["fornitore"] = {"$regex": fornitore, "$options": "i"}
    
    if tipo:
        query["tipo"] = tipo
    
    # Count totale
    total = await db["estratto_conto_movimenti"].count_documents(query)
    
    # Recupera movimenti
    movimenti = await db["estratto_conto_movimenti"].find(
        query,
        {"_id": 0}
    ).sort("data", -1).skip(offset).limit(limit).to_list(limit)
    
    # Calcola totali
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": None,
            "totale_entrate": {"$sum": {"$cond": [{"$gt": ["$importo", 0]}, "$importo", 0]}},
            "totale_uscite": {"$sum": {"$cond": [{"$lt": ["$importo", 0]}, {"$abs": "$importo"}, 0]}}
        }}
    ]
    totali_result = await db["estratto_conto_movimenti"].aggregate(pipeline).to_list(1)
    totali = totali_result[0] if totali_result else {"totale_entrate": 0, "totale_uscite": 0}
    
    return {
        "movimenti": movimenti,
        "totale": total,
        "offset": offset,
        "limit": limit,
        "totale_entrate": round(totali.get("totale_entrate", 0), 2),
        "totale_uscite": round(totali.get("totale_uscite", 0), 2)
    }


@router.get("/categorie")
@handle_errors
async def get_categorie() -> List[str]:
    """Restituisce lista categorie uniche."""
    db = Database.get_db()
    categorie = await db["estratto_conto_movimenti"].distinct("categoria")
    return sorted([c for c in categorie if c])


@router.get("/fornitori")
@handle_errors
async def get_fornitori_unici() -> List[str]:
    """Restituisce lista fornitori unici."""
    db = Database.get_db()
    fornitori = await db["estratto_conto_movimenti"].distinct("fornitore")
    return sorted([f for f in fornitori if f])


@router.get("/riepilogo")
@handle_errors
async def get_riepilogo(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    categoria: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None),
    fornitore: Optional[str] = Query(None)
) -> Dict[str, Any]:
    """Riepilogo estratto conto con filtri."""
    db = Database.get_db()
    
    query = {}
    if anno:
        if mese:
            query["data"] = {"$regex": f"^{anno}-{mese:02d}"}
        else:
            query["data"] = {"$regex": f"^{anno}"}
    
    if categoria:
        query["categoria"] = {"$regex": categoria, "$options": "i"}
    
    if tipo:
        query["tipo"] = tipo
    
    if fornitore:
        query["fornitore"] = {"$regex": fornitore, "$options": "i"}
    
    total = await db["estratto_conto_movimenti"].count_documents(query)
    
    # Totali per tipo
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$tipo",
            "totale": {"$sum": {"$abs": "$importo"}},
            "count": {"$sum": 1}
        }}
    ]
    by_tipo = await db["estratto_conto_movimenti"].aggregate(pipeline).to_list(10)
    
    entrate = next((t for t in by_tipo if t["_id"] == "entrata"), {"totale": 0, "count": 0})
    uscite = next((t for t in by_tipo if t["_id"] == "uscita"), {"totale": 0, "count": 0})
    
    # Movimenti per categoria (top 10)
    pipeline_cat = [
        {"$match": query},
        {"$group": {
            "_id": "$categoria",
            "totale": {"$sum": {"$abs": "$importo"}},
            "count": {"$sum": 1}
        }},
        {"$sort": {"totale": -1}},
        {"$limit": 10}
    ]
    by_categoria = await db["estratto_conto_movimenti"].aggregate(pipeline_cat).to_list(10)
    
    return {
        "totale_movimenti": total,
        "entrate": {"count": entrate["count"], "totale": round(entrate["totale"], 2)},
        "uscite": {"count": uscite["count"], "totale": round(uscite["totale"], 2)},
        "saldo": round(entrate["totale"] - uscite["totale"], 2),
        "per_categoria": [{"categoria": c["_id"] or "N/D", "totale": round(c["totale"], 2), "count": c["count"]} for c in by_categoria]
    }


@router.delete("/clear")
@handle_errors
async def clear_estratto_conto(anno: Optional[int] = Query(None)) -> Dict[str, Any]:
    """Elimina movimenti estratto conto."""
    db = Database.get_db()
    
    query = {}
    if anno:
        query["data"] = {"$regex": f"^{anno}"}
    
    result = await db["estratto_conto_movimenti"].delete_many(query)
    
    return {"message": f"Eliminati {result.deleted_count} movimenti"}


@router.delete("/{movimento_id}")
@handle_errors
async def elimina_singolo_movimento(movimento_id: str) -> Dict[str, Any]:
    """Elimina un singolo movimento dall'estratto conto."""
    db = Database.get_db()
    
    # Verifica che esista
    movimento = await db["estratto_conto_movimenti"].find_one({"id": movimento_id})
    if not movimento:
        raise HTTPException(status_code=404, detail="Movimento non trovato")
    
    result = await db["estratto_conto_movimenti"].delete_one({"id": movimento_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=500, detail="Errore durante l'eliminazione")
    
    return {"success": True, "message": "Movimento eliminato", "deleted_id": movimento_id}


@router.get("/export-excel")
@handle_errors
async def export_estratto_conto_excel(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    categoria: Optional[str] = Query(None),
    fornitore: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None)
):
    """
    Esporta i movimenti dell'estratto conto in formato Excel.
    Applica gli stessi filtri della visualizzazione.
    """
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    db = Database.get_db()
    
    # Costruisci query con filtri
    query = {}
    
    if anno:
        query["data"] = {"$regex": f"^{anno}"}
        if mese:
            query["data"] = {"$regex": f"^{anno}-{mese:02d}"}
    
    if categoria:
        query["categoria"] = {"$regex": categoria, "$options": "i"}
    
    if fornitore:
        query["fornitore"] = {"$regex": fornitore, "$options": "i"}
    
    if tipo:
        query["tipo"] = tipo
    
    # Recupera tutti i movimenti (senza paginazione per export)
    movimenti = await db["estratto_conto_movimenti"].find(
        query,
        {"_id": 0}
    ).sort("data", -1).to_list(10000)
    
    # Crea workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estratto Conto"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Data", "Fornitore", "Importo (€)", "Tipo", "N. Fattura", "Data Pag.", "Categoria"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Larghezze colonne
    col_widths = [12, 35, 15, 10, 30, 12, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Dati
    entrata_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    uscita_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    totale_entrate = 0
    totale_uscite = 0
    
    for row_num, mov in enumerate(movimenti, 2):
        # Formatta data
        data_str = mov.get("data", "")
        if data_str:
            try:
                parts = data_str.split("-")
                data_formatted = f"{parts[2]}/{parts[1]}/{parts[0]}"
            except (ValueError, TypeError, IndexError):
                data_formatted = data_str
        else:
            data_formatted = ""
        
        data_pag = mov.get("data_pagamento", "")
        if data_pag:
            try:
                parts = data_pag.split("-")
                data_pag_formatted = f"{parts[2]}/{parts[1]}/{parts[0]}"
            except (ValueError, TypeError, IndexError):
                data_pag_formatted = data_pag
        else:
            data_pag_formatted = ""
        
        importo = mov.get("importo", 0)
        tipo_mov = "Entrata" if importo >= 0 else "Uscita"
        
        if importo >= 0:
            totale_entrate += importo
        else:
            totale_uscite += abs(importo)
        
        row_data = [
            data_formatted,
            mov.get("fornitore") or "",
            abs(importo),
            tipo_mov,
            mov.get("numero_fattura") or "",
            data_pag_formatted,
            mov.get("categoria") or ""
        ]
        
        row_fill = entrata_fill if importo >= 0 else uscita_fill
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.fill = row_fill
            
            # Formato numerico per importo
            if col == 3:
                cell.number_format = '#,##0.00'
    
    # Riga totali
    last_row = len(movimenti) + 2
    totals_row = last_row + 1
    
    ws.cell(row=totals_row, column=1, value="TOTALI")
    ws.cell(row=totals_row, column=1).font = Font(bold=True)
    
    ws.cell(row=totals_row, column=2, value=f"Entrate: € {totale_entrate:,.2f}")
    ws.cell(row=totals_row, column=2).font = Font(bold=True, color="16A34A")
    
    ws.cell(row=totals_row, column=3, value=f"Uscite: € {totale_uscite:,.2f}")
    ws.cell(row=totals_row, column=3).font = Font(bold=True, color="DC2626")
    
    saldo = totale_entrate - totale_uscite
    ws.cell(row=totals_row, column=4, value=f"Saldo: € {saldo:,.2f}")
    ws.cell(row=totals_row, column=4).font = Font(bold=True, color="16A34A" if saldo >= 0 else "DC2626")
    
    # Salva in memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nome file
    filename_parts = ["estratto_conto"]
    if anno:
        filename_parts.append(str(anno))
    if mese:
        mesi_nomi = ['gen', 'feb', 'mar', 'apr', 'mag', 'giu', 'lug', 'ago', 'set', 'ott', 'nov', 'dic']
        filename_parts.append(mesi_nomi[mese - 1])
    filename = "_".join(filename_parts) + ".xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# ==================== RICONCILIAZIONE STIPENDI ====================

@router.post("/riconcilia-stipendi")
@handle_errors
async def riconcilia_stipendi_automatico(anno: Optional[int] = Query(None)) -> Dict[str, Any]:
    """
    Riconcilia automaticamente i bonifici stipendio con la prima nota salari.
    Cerca i movimenti "VOSTRA DISPOSIZIONE" con nomi di dipendenti e li collega.
    """
    db = Database.get_db()
    
    # Carica nomi dipendenti dalla prima_nota_salari (fonte più affidabile)
    nomi_dipendenti = await db["prima_nota_salari"].distinct("dipendente")
    dipendenti_map = {}
    for nome in nomi_dipendenti:
        if nome:
            nome_upper = nome.strip().upper()
            dipendenti_map[nome_upper] = {"nome": nome}
            # Aggiungi anche versione invertita (Nome Cognome)
            parts = nome_upper.split()
            if len(parts) >= 2:
                nome_inv = " ".join(reversed(parts))
                dipendenti_map[nome_inv] = {"nome": nome}
            # Aggiungi anche singoli cognomi/nomi (per match parziale)
            for p in parts:
                if len(p) > 3:  # Solo parole significative
                    dipendenti_map[p] = {"nome": nome}
    
    # Se vuota, prova con collezione employees (nome corretto)
    if not dipendenti_map:
        dipendenti = await db["dipendenti"].find({}, {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "nome_completo": 1}).to_list(1000)
        for d in dipendenti:
            nome_completo = d.get("nome_completo") or f"{d.get('cognome', '')} {d.get('nome', '')}".strip().upper()
            nome_inv = f"{d.get('nome', '')} {d.get('cognome', '')}".strip().upper()
            if nome_completo:
                dipendenti_map[nome_completo.upper()] = d
            if nome_inv:
                dipendenti_map[nome_inv] = d
    
    logger.info(f"Caricati {len(dipendenti_map)} nomi dipendenti")
    
    # Cerca i bonifici "VOSTRA DISPOSIZIONE" con nome dipendente non ancora riconciliati
    query = {
        "descrizione": {"$regex": "VOSTRA DISPOSIZIONE.*FAVORE|FAVORE.*", "$options": "i"},
        "$or": [
            {"riconciliato_salario": {"$exists": False}},
            {"riconciliato_salario": False}
        ]
    }
    if anno:
        query["data"] = {"$regex": f"^{anno}"}
    
    bonifici = await db["estratto_conto"].find(query, {"_id": 0}).to_list(5000)
    
    riconciliati = 0
    non_trovati = []
    
    for bonifico in bonifici:
        desc = bonifico.get("descrizione", "").upper()
        importo = abs(bonifico.get("importo", 0))
        data = bonifico.get("data", "")
        
        # Estrai nome dopo "FAVORE"
        nome_trovato = None
        dip_trovato = None
        
        if "FAVORE" in desc:
            idx = desc.find("FAVORE")
            after = desc[idx + 7:].strip()
            # Pulisci fino a separatori comuni
            for sep in [" - ", " ADD", " NOTPROVIDE", "/", "COMM."]:
                if sep in after:
                    after = after[:after.find(sep)].strip()
                    break
            
            after_clean = after.strip()
            
            # Cerca il dipendente con vari metodi
            # 1. Match esatto
            if after_clean in dipendenti_map:
                dip_trovato = dipendenti_map[after_clean]
                nome_trovato = after_clean
            else:
                # 2. Match per singola parola del nome estratto (cerca cognome)
                after_parts = [p.strip() for p in after_clean.split() if len(p.strip()) > 2]
                for part in after_parts:
                    if part in dipendenti_map:
                        dip_trovato = dipendenti_map[part]
                        nome_trovato = part
                        break
                
                # 3. Match parziale se non trovato
                if not dip_trovato:
                    for nome, dip in dipendenti_map.items():
                        if nome in after_clean or after_clean in nome:
                            dip_trovato = dip
                            nome_trovato = nome
                            break
                        # Match per cognome
                        parts = nome.split()
                        for p in parts:
                            if len(p) > 3 and p in after_clean:
                                dip_trovato = dip
                                nome_trovato = nome
                                break
                        if dip_trovato:
                            break
        
        if dip_trovato:
            # Aggiorna l'estratto conto
            await db["estratto_conto"].update_one(
                {"id": bonifico.get("id")},
                {"$set": {
                    "riconciliato_salario": True,
                    "dipendente_nome": dip_trovato.get("nome", nome_trovato),
                    "categoria": "Stipendi",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            # Cerca/aggiorna prima nota salari corrispondente
            mese_bonifico = int(data[5:7]) if len(data) >= 7 else None
            anno_bonifico = int(data[:4]) if len(data) >= 4 else None
            
            if mese_bonifico and anno_bonifico:
                nome_orig = dip_trovato.get("nome", "")
                cognome_search = nome_orig.split()[0] if nome_orig else ""
                
                if cognome_search:
                    pns = await db["prima_nota_salari"].find_one({
                        "dipendente": {"$regex": cognome_search, "$options": "i"},
                        "anno": anno_bonifico,
                        "mese": mese_bonifico,
                        "tipo": "bonifico"
                    })
                    
                    if pns:
                        await db["prima_nota_salari"].update_one(
                            {"id": pns.get("id")},
                            {"$set": {
                                "riconciliato": True,
                                "estratto_conto_id": bonifico.get("id"),
                                "importo_bonifico": importo,
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }}
                        )
            
            riconciliati += 1
        else:
            non_trovati.append({
                "data": data,
                "importo": importo,
                "descrizione": bonifico.get("descrizione", "")[:80]
            })
    
    return {
        "success": True,
        "riconciliati": riconciliati,
        "non_trovati_count": len(non_trovati),
        "non_trovati_sample": non_trovati[:10],
        "message": f"Riconciliati {riconciliati} bonifici stipendio"
    }


@router.get("/movimenti-stipendi")
@handle_errors
async def get_movimenti_stipendi(
    anno: Optional[int] = Query(None),
    solo_non_riconciliati: bool = Query(False)
) -> Dict[str, Any]:
    """
    Restituisce i movimenti dell'estratto conto che sembrano essere stipendi.
    """
    db = Database.get_db()
    
    query = {
        "descrizione": {"$regex": "VOSTRA DISPOSIZIONE|VS\\.DISP", "$options": "i"},
        "tipo": "uscita"
    }
    
    if anno:
        query["data"] = {"$regex": f"^{anno}"}
    
    if solo_non_riconciliati:
        query["$or"] = [
            {"riconciliato_salario": {"$exists": False}},
            {"riconciliato_salario": False}
        ]
    
    movimenti = await db["estratto_conto"].find(query, {"_id": 0}).sort("data", -1).to_list(1000)
    
    # Raggruppa per dipendente se riconciliato
    per_dipendente = {}
    non_riconciliati = []
    
    for m in movimenti:
        if m.get("riconciliato_salario") and m.get("dipendente_nome"):
            nome = m.get("dipendente_nome")
            if nome not in per_dipendente:
                per_dipendente[nome] = {"count": 0, "totale": 0, "movimenti": []}
            per_dipendente[nome]["count"] += 1
            per_dipendente[nome]["totale"] += abs(m.get("importo", 0))
            if len(per_dipendente[nome]["movimenti"]) < 5:
                per_dipendente[nome]["movimenti"].append(m)
        else:
            non_riconciliati.append(m)
    
    return {
        "totale": len(movimenti),
        "riconciliati": len(movimenti) - len(non_riconciliati),
        "non_riconciliati": len(non_riconciliati),
        "per_dipendente": [
            {"nome": k, **v} for k, v in sorted(per_dipendente.items(), key=lambda x: x[1]["totale"], reverse=True)
        ],
        "non_riconciliati_sample": non_riconciliati[:20]
    }

