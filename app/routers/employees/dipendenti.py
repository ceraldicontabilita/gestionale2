"""
Gestione Dipendenti - Router API completo.
Anagrafica, turni, libro unico, libretti sanitari.
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Body
from typing import Dict, Any, List, Optional
from datetime import datetime, timedelta, timezone
import uuid
import logging
import io
import re

from app.database import Database, Collections
from app.utils.error_handler import handle_errors

logger = logging.getLogger(__name__)
router = APIRouter()

# Costanti
TURNI_TIPI = {
    "mattina": {"label": "Mattina", "orario": "06:00 - 14:00", "color": "#4caf50"},
    "pomeriggio": {"label": "Pomeriggio", "orario": "14:00 - 22:00", "color": "#2196f3"},
    "sera": {"label": "Sera", "orario": "18:00 - 02:00", "color": "#9c27b0"},
    "full": {"label": "Full Day", "orario": "10:00 - 22:00", "color": "#ff9800"},
    "riposo": {"label": "Riposo", "orario": "-", "color": "#9e9e9e"},
    "ferie": {"label": "Ferie", "orario": "-", "color": "#e91e63"},
    "malattia": {"label": "Malattia", "orario": "-", "color": "#f44336"}
}

MANSIONI = [
    "Cameriere", "Cuoco", "Aiuto Cuoco", "Barista", "Pizzaiolo", 
    "Lavapiatti", "Cassiera", "Responsabile Sala", "Chef", "Sommelier"
]

CONTRATTI_TIPI = [
    "Tempo Indeterminato", "Tempo Determinato", "Apprendistato", 
    "Stage/Tirocinio", "Collaborazione", "Part-time"
]


@router.get("")
@handle_errors
async def list_dipendenti(
    skip: int = Query(0, ge=0),
    limit: int = Query(1000, ge=1, le=10000),
    attivo: Optional[bool] = Query(None),
    in_carico: Optional[bool] = Query(None, description="Se true solo in carico, false solo non in carico, null=tutti"),
    include_merged: bool = Query(False, description="Includi record già unificati (merged_into)"),
    mansione: Optional[str] = Query(None),
    search: Optional[str] = Query(None)
) -> List[Dict[str, Any]]:
    """Lista dipendenti con filtri."""
    db = Database.get_db()
    
    query: Dict[str, Any] = {}
    if attivo is not None:
        query["attivo"] = attivo
    if in_carico is True:
        # "in carico" = flag esplicito true OPPURE assente (default storico)
        query["$and"] = query.get("$and", []) + [{"$or": [{"in_carico": True}, {"in_carico": {"$exists": False}}]}]
    elif in_carico is False:
        query["in_carico"] = False
    if not include_merged:
        query["merged_into"] = {"$exists": False}
    if mansione:
        query["mansione"] = mansione
    if search:
        import re as _re
        safe_search = _re.escape(search)
        query["$or"] = [
            {"nome_completo": {"$regex": safe_search, "$options": "i"}},
            {"codice_fiscale": {"$regex": safe_search, "$options": "i"}}
        ]
    
    dipendenti_raw = await db[Collections.EMPLOYEES].find(query, {"_id": 0}).sort("nome_completo", 1).skip(skip).limit(limit).to_list(limit)
    
    # Deduplicazione per codice fiscale (prevenzione duplicati accidentali)
    seen_cf = set()
    dipendenti = []
    for d in dipendenti_raw:
        cf_key = (d.get("codice_fiscale") or "").upper().strip()
        if cf_key and cf_key in seen_cf:
            continue
        if cf_key:
            seen_cf.add(cf_key)
        dipendenti.append(d)
    
    return dipendenti


@router.get("/by-google-email")
async def get_dipendente_by_google_email(email: str = Query(...)):
    """Cerca dipendente associato a un Google email (per portale)."""
    from fastapi import HTTPException
    db = Database.get_db()
    dip = await db[Collections.EMPLOYEES].find_one(
        {"google_email": email.lower().strip()},
        {"_id": 0, "nome_completo": 1, "mansione": 1, "data_inizio_contratto": 1, "id": 1}
    )
    if not dip:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    return dip




@router.get("/stats")
@handle_errors
async def get_dipendenti_stats() -> Dict[str, Any]:
    """Statistiche dipendenti."""
    db = Database.get_db()
    
    total = await db[Collections.EMPLOYEES].count_documents({})
    attivi = await db[Collections.EMPLOYEES].count_documents({"attivo": {"$ne": False}})
    
    # Per mansione
    pipeline = [
        {"$group": {"_id": "$mansione", "count": {"$sum": 1}}}
    ]
    by_mansione = await db[Collections.EMPLOYEES].aggregate(pipeline).to_list(100)
    
    # Libretti in scadenza (prossimi 30 giorni)
    today = datetime.now(timezone.utc)
    deadline = today + timedelta(days=30)
    libretti_scadenza = await db[Collections.EMPLOYEES].count_documents({
        "libretto_scadenza": {"$lte": deadline.isoformat()[:10], "$gte": today.isoformat()[:10]}
    })
    
    return {
        "totale": total,
        "attivi": attivi,
        "inattivi": total - attivi,
        "per_mansione": {item["_id"] or "N/D": item["count"] for item in by_mansione},
        "libretti_in_scadenza": libretti_scadenza
    }


# ─── Deduplica dipendenti ────────────────────────────────────────────────────

@router.get("/duplicati")
@handle_errors
async def lista_duplicati_dipendenti() -> Dict[str, Any]:
    """
    Analizza l'anagrafica e ritorna gruppi di sospetti duplicati.
    Si basa su CF normalizzato e su nome+cognome normalizzati.
    """
    from app.services.dipendenti_dedupe import trova_duplicati
    return await trova_duplicati()


@router.post("/duplicati/merge")
@handle_errors
async def merge_duplicato_dipendente(
    payload: Dict[str, Any] = Body(...)
) -> Dict[str, Any]:
    """
    Unifica `duplicate_id` dentro `target_id`. Re-point di tutti i cedolini,
    presenze, verbali, movimenti. I cedolini duplicati (stesso anno+mese)
    vengono scartati. Soft delete di default (il duplicato resta come
    `merged_into` al target e `in_carico=False`).
    """
    from app.services.dipendenti_dedupe import merge_dipendenti
    target_id = payload.get("target_id")
    duplicate_id = payload.get("duplicate_id")
    soft = payload.get("soft", True)
    if not target_id or not duplicate_id:
        raise HTTPException(status_code=400, detail="target_id e duplicate_id richiesti")
    try:
        return await merge_dipendenti(target_id, duplicate_id, soft=soft)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/duplicati/auto-merge")
@handle_errors
async def auto_merge_duplicati(
    payload: Dict[str, Any] = Body(default={})
) -> Dict[str, Any]:
    """
    Auto-merge di tutti i duplicati ad alta certezza.
    Default dry_run=True: ritorna lista dei merge previsti senza eseguirli.
    Passa `{"dry_run": false}` per eseguire effettivamente i merge.
    """
    from app.services.dipendenti_dedupe import auto_merge_tutti
    dry_run = bool(payload.get("dry_run", True))
    return await auto_merge_tutti(dry_run=dry_run)


@router.get("/report-ferie-permessi-tutti")
@handle_errors
async def genera_report_ferie_permessi_tutti(
    anno: int = Query(None, description="Anno di riferimento")
):
    """
    Genera report riepilogativo ferie/permessi per TUTTI i dipendenti.
    Restituisce un PDF con tabella riassuntiva.
    """
    from fastapi.responses import Response
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER
    
    db = Database.get_db()
    
    if not anno:
        anno = datetime.now().year
    
    # Recupera tutti i dipendenti con progressivi
    dipendenti_raw = await db[Collections.EMPLOYEES].find(
        {"progressivi": {"$exists": True}},
        {"_id": 0, "nome_completo": 1, "name": 1, "cognome": 1, "nome": 1, "codice_fiscale": 1, "progressivi": 1, "attivo": 1}
    ).sort("nome_completo", 1).to_list(100)
    
    # Deduplicazione per codice fiscale (prevenzione duplicati)
    seen_cf = set()
    dipendenti = []
    for d in dipendenti_raw:
        cf_key = d.get("codice_fiscale", "").upper().strip()
        nome_key = (d.get("nome_completo") or d.get("name") or "").lower().strip()
        dedup_key = cf_key if cf_key else nome_key
        if dedup_key and dedup_key not in seen_cf:
            seen_cf.add(dedup_key)
            dipendenti.append(d)
        elif not dedup_key:
            dipendenti.append(d)
    
    if not dipendenti:
        raise HTTPException(status_code=404, detail="Nessun dipendente con progressivi trovato")
    
    # Genera PDF landscape
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=15)
    
    elements = []
    
    # Intestazione
    elements.append(Paragraph(f"RIEPILOGO FERIE E PERMESSI - ANNO {anno}", title_style))
    elements.append(Paragraph(f"Data: {datetime.now().strftime('%d/%m/%Y')}", 
                              ParagraphStyle('Date', parent=styles['Normal'], alignment=TA_CENTER)))
    elements.append(Spacer(1, 15))
    
    # Tabella
    data = [["Dipendente", "CF", "Ferie Mat.", "Ferie God.", "Ferie Res.", "Perm. Mat.", "Perm. God.", "Perm. Res.", "Stato"]]
    
    totali = {"ferie_mat": 0, "ferie_god": 0, "ferie_res": 0, "perm_mat": 0, "perm_god": 0, "perm_res": 0}
    
    for dip in dipendenti:
        nome = dip.get("nome_completo") or dip.get("name") or f"{dip.get('cognome', '')} {dip.get('nome', '')}"
        cf = dip.get("codice_fiscale", "")[:6] + "..." if dip.get("codice_fiscale") else ""
        prog = dip.get("progressivi", {})
        
        ferie_mat = prog.get("ferie_maturate", 0) or 0
        ferie_god = prog.get("ferie_godute", 0) or 0
        ferie_res = prog.get("ferie_residue", 0) or 0
        perm_mat = prog.get("permessi_maturati", 0) or 0
        perm_god = prog.get("permessi_goduti", 0) or 0
        perm_res = prog.get("permessi_residui", 0) or 0
        stato = "Attivo" if dip.get("attivo", True) else "Inattivo"
        
        totali["ferie_mat"] += ferie_mat
        totali["ferie_god"] += ferie_god
        totali["ferie_res"] += ferie_res
        totali["perm_mat"] += perm_mat
        totali["perm_god"] += perm_god
        totali["perm_res"] += perm_res
        
        data.append([
            nome[:25],
            cf,
            f"{ferie_mat:.1f}",
            f"{ferie_god:.1f}",
            f"{ferie_res:.1f}",
            f"{perm_mat:.0f}",
            f"{perm_god:.0f}",
            f"{perm_res:.0f}",
            stato
        ])
    
    # Riga totali
    data.append([
        "TOTALE",
        "",
        f"{totali['ferie_mat']:.1f}",
        f"{totali['ferie_god']:.1f}",
        f"{totali['ferie_res']:.1f}",
        f"{totali['perm_mat']:.0f}",
        f"{totali['perm_god']:.0f}",
        f"{totali['perm_res']:.0f}",
        f"{len(dipendenti)} dip."
    ])
    
    col_widths = [5*cm, 2.5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm]
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    filename = f"Riepilogo_Ferie_Permessi_{anno}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/sync-iban")
@handle_errors
async def sync_iban_field() -> Dict[str, Any]:
    """
    Sincronizza il campo 'iban' (singolo) con 'ibans' (array).
    Per ogni dipendente con 'iban' popolato ma 'ibans' vuoto,
    copia il valore in un array.
    """
    db = Database.get_db()
    
    # Trova dipendenti con iban ma senza ibans
    dipendenti = await db[Collections.EMPLOYEES].find(
        {
            "iban": {"$exists": True, "$nin": [None, ""]},
            "$or": [
                {"ibans": {"$exists": False}},
                {"ibans": None},
                {"ibans": []},
                {"ibans": {"$size": 0}}
            ]
        },
        {"id": 1, "iban": 1, "_id": 0}
    ).to_list(1000)
    
    aggiornati = 0
    for dip in dipendenti:
        iban = dip.get("iban", "").upper().replace(" ", "")
        if iban:
            await db[Collections.EMPLOYEES].update_one(
                {"id": dip["id"]},
                {"$set": {"ibans": [iban]}}
            )
            aggiornati += 1
    
    return {
        "success": True,
        "dipendenti_analizzati": len(dipendenti),
        "dipendenti_aggiornati": aggiornati
    }


@router.get("/tipi-turno")
@handle_errors
async def get_tipi_turno() -> Dict[str, Any]:
    """Ritorna i tipi di turno disponibili."""
    return TURNI_TIPI


@router.get("/mansioni")
@handle_errors
async def get_mansioni() -> List[str]:
    """Ritorna le mansioni disponibili."""
    return MANSIONI


@router.get("/tipi-contratto")
@handle_errors
async def get_tipi_contratto() -> List[str]:
    """Ritorna i tipi di contratto disponibili."""
    return CONTRATTI_TIPI


@router.post("")
@handle_errors
async def create_dipendente(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Crea nuovo dipendente."""
    db = Database.get_db()
    
    # Campi obbligatori
    required = ["nome_completo"]
    for field in required:
        if not data.get(field):
            raise HTTPException(status_code=400, detail=f"Campo obbligatorio mancante: {field}")
    
    # Parse nome
    nome_parts = data["nome_completo"].split()
    cognome = nome_parts[0] if nome_parts else ""
    nome = " ".join(nome_parts[1:]) if len(nome_parts) > 1 else ""
    
    dipendente = {
        "id": str(uuid.uuid4()),
        "nome_completo": data["nome_completo"],
        "cognome": cognome,
        "nome": nome,
        "codice_fiscale": data.get("codice_fiscale", ""),
        "codice_dipendente": data.get("codice_dipendente", ""),  # Codice aziendale (es. 0300006)
        "matricola": data.get("matricola", ""),
        "email": data.get("email", ""),
        "telefono": data.get("telefono", ""),
        "indirizzo": data.get("indirizzo", ""),
        "data_nascita": data.get("data_nascita"),
        "luogo_nascita": data.get("luogo_nascita", ""),
        "mansione": data.get("mansione", ""),  # es. "CAM. DI SALA"
        "qualifica": data.get("qualifica", ""),  # es. "OPE" (operatore)
        "livello": data.get("livello", ""),  # es. "6 Livello Super"
        "tipo_contratto": data.get("tipo_contratto", "Tempo Indeterminato"),
        "data_assunzione": data.get("data_assunzione"),
        "data_fine_contratto": data.get("data_fine_contratto"),
        "ore_settimanali": data.get("ore_settimanali", 40),
        "giorni_lavoro": data.get("giorni_lavoro", ["lun", "mar", "mer", "gio", "ven", "sab"]),  # Default Lun-Sab
        "iban": data.get("iban", ""),
        "ibans": data.get("ibans", []),  # IBAN multipli (principale + secondari)
        # Retribuzione dettagliata
        "paga_base": data.get("paga_base", 0),  # Paga base mensile
        "contingenza": data.get("contingenza", 0),  # Indennità contingenza
        "stipendio_lordo": data.get("stipendio_lordo", 0),  # Totale lordo
        "stipendio_orario": data.get("stipendio_orario", 0),  # Paga oraria
        # Agevolazioni fiscali
        "agevolazioni": data.get("agevolazioni", []),  # es. ["Decontr.SUD DL104.20"]
        # Progressivi
        "progressivi": {
            "tfr_accantonato": data.get("tfr_accantonato", 0),
            "ferie_maturate": data.get("ferie_maturate", 0),
            "ferie_godute": data.get("ferie_godute", 0),
            "ferie_residue": data.get("ferie_residue", 0),
            "permessi_maturati": data.get("permessi_maturati", 0),
            "permessi_goduti": data.get("permessi_goduti", 0),
            "permessi_residui": data.get("permessi_residui", 0),
            "rol_maturati": data.get("rol_maturati", 0),
            "rol_goduti": data.get("rol_goduti", 0),
            "rol_residui": data.get("rol_residui", 0)
        },
        # Acconti
        "acconti": data.get("acconti", []),
        # Libretto sanitario
        "libretto_numero": data.get("libretto_numero", ""),
        "libretto_scadenza": data.get("libretto_scadenza"),
        "libretto_file": data.get("libretto_file"),
        # Portale
        "portale_invitato": False,
        "portale_registrato": False,
        "portale_ultimo_accesso": None,
        # Bonifici associati
        "bonifici_associati": data.get("bonifici_associati", []),
        # Status
        "attivo": True,
        "in_carico": data.get("in_carico", True),  # Flag per presenze (default: true per nuovi dipendenti)
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Verifica duplicato CF
    if dipendente["codice_fiscale"]:
        existing = await db[Collections.EMPLOYEES].find_one({"codice_fiscale": dipendente["codice_fiscale"]})
        if existing:
            raise HTTPException(status_code=409, detail="Dipendente con questo codice fiscale già esistente")
    
    await db[Collections.EMPLOYEES].insert_one(dipendente.copy())
    dipendente.pop("_id", None)
    
    return dipendente


# ============== BUSTE PAGA (must be before /{dipendente_id} to avoid route conflict) ==============

@router.get("/buste-paga")
@handle_errors
async def get_buste_paga(
    anno: int = Query(...),
    mese: str = Query(...)
) -> List[Dict[str, Any]]:
    """
    Ottiene le buste paga per un determinato mese.
    Le buste paga vengono create automaticamente dai movimenti salari.
    """
    db = Database.get_db()
    
    periodo = f"{anno}-{mese}"
    
    # Cerca buste paga esistenti
    buste = await db["cedolini"].find(
        {"periodo": periodo},
        {"_id": 0}
    ).to_list(1000)
    
    return buste


@router.post("/buste-paga")
@handle_errors
async def create_busta_paga(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Crea o aggiorna una busta paga."""
    db = Database.get_db()
    
    required = ["dipendente_id", "periodo"]
    for field in required:
        if not data.get(field):
            raise HTTPException(status_code=400, detail=f"Campo {field} obbligatorio")
    
    # Cerca busta esistente
    existing = await db["cedolini"].find_one({
        "dipendente_id": data["dipendente_id"],
        "periodo": data["periodo"]
    })
    
    busta = {
        "dipendente_id": data["dipendente_id"],
        "periodo": data["periodo"],
        "lordo": float(data.get("lordo", 0) or 0),
        "netto": float(data.get("netto", 0) or 0),
        "contributi": float(data.get("contributi", 0) or 0),
        "trattenute": float(data.get("trattenute", 0) or 0),
        "pagata": bool(data.get("pagata", False)),
        "data_pagamento": data.get("data_pagamento"),
        "note": data.get("note", ""),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if existing:
        await db["cedolini"].update_one(
            {"id": existing["id"]},
            {"$set": busta}
        )
        busta["id"] = existing["id"]
    else:
        busta["id"] = str(uuid.uuid4())
        busta["created_at"] = datetime.now(timezone.utc).isoformat()
        await db["cedolini"].insert_one(busta.copy())
    
    busta.pop("_id", None)
    return busta


# ============== NOTA: Sezione SALARI rimossa ==============
# La gestione salari/prima nota è stata spostata in /app/app/routers/prima_nota_salari.py
# Endpoints disponibili: /api/prima-nota-salari/*

# ============== GESTIONE CONTRATTI - lista (must be BEFORE /{dipendente_id}) ==============

@router.get("/contratti")
@handle_errors
async def list_contratti_proxy(
    dipendente_id: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None),
    stato: Optional[str] = Query(None)
) -> List[Dict[str, Any]]:
    """Lista tutti i contratti (proxy pre-routing per evitare conflitto con /{dipendente_id})."""
    db = Database.get_db()
    query = {}
    if dipendente_id:
        query["dipendente_id"] = dipendente_id
    if tipo:
        query["tipo_contratto"] = tipo
    if stato:
        query["stato"] = stato
    contratti = await db["contratti_dipendenti"].find(query, {"_id": 0}).sort("data_inizio", -1).to_list(500)
    return contratti


# ============== DIPENDENTE DETAIL (must be after specific routes) ==============

@router.get("/{dipendente_id}")
@handle_errors
async def get_dipendente(dipendente_id: str) -> Dict[str, Any]:
    """Dettaglio singolo dipendente."""
    db = Database.get_db()
    
    dipendente = await db[Collections.EMPLOYEES].find_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"_id": 0}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return dipendente


@router.put("/{dipendente_id}")
@handle_errors
async def update_dipendente(dipendente_id: str, data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Aggiorna dipendente e sincronizza IBAN nei bonifici associati."""
    db = Database.get_db()
    
    # Rimuovi campi non modificabili
    data.pop("id", None)
    data.pop("created_at", None)
    
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Trova il dipendente prima dell'update
    dipendente_old = await db[Collections.EMPLOYEES].find_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"_id": 0}
    )
    
    if not dipendente_old:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Aggiorna il dipendente
    result = await db[Collections.EMPLOYEES].update_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # SINCRONIZZAZIONE A CASCATA: Aggiorna IBAN nei bonifici associati
    new_ibans = data.get("ibans", [])
    new_iban = data.get("iban", "")
    dip_id = dipendente_old.get("id")
    
    if new_ibans or new_iban:
        # Aggiorna i bonifici associati a questo dipendente
        all_ibans = list(set([i for i in new_ibans if i] + ([new_iban] if new_iban else [])))
        
        # Aggiorna i bonifici dove il dipendente è già associato
        await db.bonifici_transfers.update_many(
            {"dipendente_id": dip_id},
            {"$set": {
                "dipendente_ibans": all_ibans,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Inoltre, trova bonifici con IBAN corrispondente e associali automaticamente se non già associati
        if all_ibans:
            nome_completo = data.get("nome_completo") or dipendente_old.get("nome_completo") or \
                f"{data.get('nome', dipendente_old.get('nome', ''))} {data.get('cognome', dipendente_old.get('cognome', ''))}".strip()
            
            # Cerca bonifici con IBAN beneficiario corrispondente
            for iban in all_ibans:
                if iban and len(iban) >= 15:  # IBAN minimo valido
                    await db.bonifici_transfers.update_many(
                        {
                            "beneficiario.iban": iban,
                            "dipendente_id": {"$exists": False}  # Solo se non già associato
                        },
                        {"$set": {
                            "dipendente_id": dip_id,
                            "dipendente_nome": nome_completo,
                            "dipendente_ibans": all_ibans,
                            "auto_match_iban": True,
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )
    
    return {"message": "Dipendente aggiornato"}


@router.delete("/{dipendente_id}")
@handle_errors
async def delete_dipendente(dipendente_id: str) -> Dict[str, str]:
    """Elimina dipendente."""
    db = Database.get_db()
    
    result = await db[Collections.EMPLOYEES].delete_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return {"message": "Dipendente eliminato"}


# ============== TURNI ==============

@router.get("/turni/settimana")
@handle_errors
async def get_turni_settimana(
    data_inizio: str = Query(..., description="Data inizio settimana (YYYY-MM-DD)")
) -> Dict[str, Any]:
    """Ritorna i turni per una settimana."""
    db = Database.get_db()
    
    # Calcola date settimana
    start = datetime.strptime(data_inizio, "%Y-%m-%d")
    dates = [(start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]
    
    # Trova turni
    turni = await db["turni_dipendenti"].find(
        {"data": {"$in": dates}},
        {"_id": 0}
    ).to_list(1000)
    
    # Organizza per dipendente e data
    turni_by_employee = {}
    for t in turni:
        emp_id = t.get("dipendente_id")
        if emp_id not in turni_by_employee:
            turni_by_employee[emp_id] = {}
        turni_by_employee[emp_id][t.get("data")] = t.get("turno")
    
    # Carica dipendenti attivi
    dipendenti = await db[Collections.EMPLOYEES].find(
        {"attivo": {"$ne": False}},
        {"_id": 0, "id": 1, "nome_completo": 1, "mansione": 1}
    ).to_list(100)
    
    return {
        "settimana": dates,
        "dipendenti": dipendenti,
        "turni": turni_by_employee
    }


@router.post("/turni/salva")
@handle_errors
async def salva_turni(data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Salva turni per una settimana."""
    db = Database.get_db()
    
    turni = data.get("turni", {})  # {dipendente_id: {data: turno}}
    
    for dip_id, turni_dip in turni.items():
        for data_turno, tipo_turno in turni_dip.items():
            await db["turni_dipendenti"].update_one(
                {"dipendente_id": dip_id, "data": data_turno},
                {"$set": {
                    "dipendente_id": dip_id,
                    "data": data_turno,
                    "turno": tipo_turno,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }},
                upsert=True
            )
    
    return {"message": "Turni salvati"}


# ============== LIBRETTI SANITARI ==============

@router.get("/libretti/scadenze")
@handle_errors
async def get_libretti_scadenze(days: int = Query(30, ge=1, le=365)) -> List[Dict[str, Any]]:
    """Ritorna dipendenti con libretto in scadenza."""
    db = Database.get_db()
    
    today = datetime.now(timezone.utc)
    deadline = today + timedelta(days=days)
    
    dipendenti = await db[Collections.EMPLOYEES].find(
        {
            "libretto_scadenza": {"$ne": None},
            "$or": [
                {"libretto_scadenza": {"$lte": deadline.isoformat()[:10]}},
                {"libretto_scadenza": {"$lt": today.isoformat()[:10]}}  # Già scaduti
            ]
        },
        {"_id": 0}
    ).sort("libretto_scadenza", 1).to_list(100)
    
    return dipendenti


@router.put("/{dipendente_id}/libretto")
@handle_errors
async def update_libretto(dipendente_id: str, data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Aggiorna dati libretto sanitario."""
    db = Database.get_db()
    
    update_data = {}
    if "libretto_numero" in data:
        update_data["libretto_numero"] = data["libretto_numero"]
    if "libretto_scadenza" in data:
        update_data["libretto_scadenza"] = data["libretto_scadenza"]
    if "libretto_file" in data:
        update_data["libretto_file"] = data["libretto_file"]
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db[Collections.EMPLOYEES].update_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return {"message": "Libretto aggiornato"}


# ============== PORTALE DIPENDENTI ==============

@router.post("/{dipendente_id}/invita-portale")
@handle_errors
async def invita_portale(dipendente_id: str) -> Dict[str, str]:
    """Segna dipendente come invitato al portale."""
    db = Database.get_db()
    
    result = await db[Collections.EMPLOYEES].update_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"$set": {
            "portale_invitato": True,
            "portale_data_invito": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return {"message": "Invito inviato"}


@router.post("/invita-multipli")
@handle_errors
async def invita_multipli(dipendenti_ids: List[str] = Body(...)) -> Dict[str, Any]:
    """Invita multipli dipendenti al portale."""
    db = Database.get_db()
    
    result = await db[Collections.EMPLOYEES].update_many(
        {"id": {"$in": dipendenti_ids}},
        {"$set": {
            "portale_invitato": True,
            "portale_data_invito": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Invitati {result.modified_count} dipendenti"}


# ============== LIBRETTI SANITARI - COLLECTION SEPARATA ==============

@router.get("/libretti-sanitari/all")
@handle_errors
async def get_all_libretti_sanitari() -> List[Dict[str, Any]]:
    """Lista tutti i libretti sanitari."""
    db = Database.get_db()
    
    libretti = await db["libretti_sanitari"].find({}, {"_id": 0}).sort("data_scadenza", 1).to_list(500)
    return libretti


@router.post("/libretti-sanitari")
@handle_errors
async def create_libretto_sanitario(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Crea nuovo libretto sanitario."""
    db = Database.get_db()
    
    libretto = {
        "id": str(uuid.uuid4()),
        "dipendente_nome": data.get("dipendente_nome", ""),
        "dipendente_id": data.get("dipendente_id"),
        "numero_libretto": data.get("numero_libretto", ""),
        "data_rilascio": data.get("data_rilascio"),
        "data_scadenza": data.get("data_scadenza"),
        "stato": data.get("stato", "valido"),
        "note": data.get("note", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db["libretti_sanitari"].insert_one(libretto.copy())
    libretto.pop("_id", None)
    
    return libretto


@router.put("/libretti-sanitari/{libretto_id}")
@handle_errors
async def update_libretto_sanitario(libretto_id: str, data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Aggiorna libretto sanitario."""
    db = Database.get_db()
    
    data.pop("id", None)
    data.pop("created_at", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db["libretti_sanitari"].update_one(
        {"id": libretto_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Libretto non trovato")
    
    return {"message": "Libretto aggiornato"}


@router.delete("/libretti-sanitari/{libretto_id}")
@handle_errors
async def delete_libretto_sanitario(libretto_id: str) -> Dict[str, str]:
    """Elimina libretto sanitario."""
    db = Database.get_db()
    
    result = await db["libretti_sanitari"].delete_one({"id": libretto_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Libretto non trovato")
    
    return {"message": "Libretto eliminato"}


# ============== LIBRO UNICO ==============

@router.get("/libro-unico/presenze")
@handle_errors
async def get_libro_unico_presenze(month_year: str = Query(..., description="Formato: YYYY-MM")) -> List[Dict[str, Any]]:
    """Ottieni presenze dal libro unico per un mese."""
    db = Database.get_db()
    
    presenze = await db["libro_unico_presenze"].find(
        {"month_year": month_year},
        {"_id": 0}
    ).sort("dipendente_nome", 1).to_list(500)
    
    return presenze


@router.get("/libro-unico/salaries")
@handle_errors
async def get_libro_unico_salaries(month_year: Optional[str] = Query(None, description="Formato: YYYY-MM")) -> List[Dict[str, Any]]:
    """Ottieni buste paga dal libro unico."""
    db = Database.get_db()
    
    query = {}
    if month_year:
        query["month_year"] = month_year
    
    salaries = await db["libro_unico_salaries"].find(query, {"_id": 0}).sort([("month_year", -1), ("dipendente_nome", 1)]).to_list(1000)
    
    return salaries


@router.post("/libro-unico/upload")
@handle_errors
async def upload_libro_unico(
    file: UploadFile = File(...),
    month_year: str = Query(..., description="Formato: YYYY-MM")
) -> Dict[str, Any]:
    """
    Upload e parsing di PDF/Excel del libro unico.
    Estrae presenze, buste paga e aggiorna anagrafica.
    """
    db = Database.get_db()
    
    filename = file.filename.lower()
    content = await file.read()
    
    presenze_count = 0
    salaries_count = 0
    payments_count = 0
    anagrafica_created = 0
    anagrafica_updated = 0
    
    if filename.endswith('.pdf'):
        # Parsing PDF - estrai dati buste paga
        try:
            import fitz  # PyMuPDF
            
            pdf_doc = fitz.open(stream=content, filetype="pdf")
            
            for page in pdf_doc:
                text = page.get_text()
                
                # Cerca pattern busta paga
                # Pattern: NOME COGNOME seguito da valori numerici
                lines = text.split('\n')
                
                current_employee = None
                for i, line in enumerate(lines):
                    line = line.strip()
                    
                    # Cerca nome dipendente (tutte maiuscole, senza numeri)
                    if line and line.isupper() and not any(c.isdigit() for c in line) and len(line) > 3:
                        # Potrebbe essere un nome
                        if len(line.split()) >= 2:
                            current_employee = line
                    
                    # Cerca pattern "NETTO A PAGARE" o simili
                    if current_employee and 'netto' in line.lower():
                        # Cerca importo nelle righe successive
                        for j in range(i, min(i+5, len(lines))):
                            next_line = lines[j].strip()
                            # Cerca pattern €1.234,56 o 1234.56
                            import re
                            amounts = re.findall(r'[\d.,]+', next_line)
                            for amt in amounts:
                                try:
                                    # Converti formato italiano
                                    amt_clean = amt.replace('.', '').replace(',', '.')
                                    netto = float(amt_clean)
                                    if 100 < netto < 10000:  # Range plausibile per stipendio
                                        # Salva busta paga
                                        salary_doc = {
                                            "id": str(uuid.uuid4()),
                                            "dipendente_nome": current_employee,
                                            "month_year": month_year,
                                            "netto_a_pagare": netto,
                                            "acconto_pagato": 0,
                                            "differenza": netto,
                                            "note": f"Importato da PDF: {file.filename}",
                                            "created_at": datetime.now(timezone.utc).isoformat()
                                        }
                                        
                                        # Verifica se esiste già
                                        existing = await db["libro_unico_salaries"].find_one({
                                            "dipendente_nome": current_employee,
                                            "month_year": month_year
                                        })
                                        
                                        if not existing:
                                            await db["libro_unico_salaries"].insert_one(salary_doc.copy())
                                            salaries_count += 1
                                        
                                        current_employee = None
                                        break
                                except (ValueError, TypeError):
                                    continue
                            if current_employee is None:
                                break
            
            pdf_doc.close()
            
        except ImportError:
            raise HTTPException(status_code=500, detail="PyMuPDF non installato per parsing PDF")
        except Exception as e:
            logger.error(f"Errore parsing PDF: {e}")
            raise HTTPException(status_code=400, detail=f"Errore parsing PDF: {str(e)}")
    
    elif filename.endswith(('.xlsx', '.xls')):
        # Parsing Excel
        try:
            import pandas as pd
            
            df = pd.read_excel(io.BytesIO(content))
            
            # Cerca colonne rilevanti
            columns_lower = {c.lower(): c for c in df.columns}
            
            for idx, row in df.iterrows():
                try:
                    # Cerca nome dipendente
                    nome = None
                    for col_key in ['nome', 'dipendente', 'cognome e nome', 'nominativo']:
                        if col_key in columns_lower:
                            nome = str(row[columns_lower[col_key]]) if pd.notna(row[columns_lower[col_key]]) else None
                            break
                    
                    if not nome or nome == 'nan':
                        continue
                    
                    # Cerca importo netto
                    netto = None
                    for col_key in ['netto', 'netto a pagare', 'importo', 'stipendio']:
                        if col_key in columns_lower:
                            val = row[columns_lower[col_key]]
                            if pd.notna(val):
                                try:
                                    netto = float(val)
                                except (ValueError, TypeError):
                                    pass
                            break
                    
                    if netto and netto > 0:
                        salary_doc = {
                            "id": str(uuid.uuid4()),
                            "dipendente_nome": nome.upper(),
                            "month_year": month_year,
                            "netto_a_pagare": netto,
                            "acconto_pagato": 0,
                            "differenza": netto,
                            "note": f"Importato da Excel: {file.filename}",
                            "created_at": datetime.now(timezone.utc).isoformat()
                        }
                        
                        existing = await db["libro_unico_salaries"].find_one({
                            "dipendente_nome": nome.upper(),
                            "month_year": month_year
                        })
                        
                        if not existing:
                            await db["libro_unico_salaries"].insert_one(salary_doc.copy())
                            salaries_count += 1
                        
                except Exception as e:
                    logger.warning(f"Errore riga {idx}: {e}")
                    continue
                    
        except Exception as e:
            logger.error(f"Errore parsing Excel: {e}")
            raise HTTPException(status_code=400, detail=f"Errore parsing Excel: {str(e)}")
    
    else:
        raise HTTPException(status_code=400, detail="Formato non supportato. Usa PDF o Excel.")
    
    return {
        "success": True,
        "message": f"Import completato per {month_year}",
        "presenze_count": presenze_count,
        "salaries_count": salaries_count,
        "payments_count": payments_count,
        "anagrafica_created": anagrafica_created,
        "anagrafica_updated": anagrafica_updated
    }


@router.get("/libro-unico/export-excel")
@handle_errors
async def export_libro_unico_excel(month_year: str = Query(..., description="Formato: YYYY-MM")):
    """Esporta libro unico in Excel."""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    db = Database.get_db()
    
    # Recupera dati
    salaries = await db["libro_unico_salaries"].find(
        {"month_year": month_year},
        {"_id": 0}
    ).sort("dipendente_nome", 1).to_list(500)
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Libro Unico {month_year}"
    
    # Header
    headers = ["Dipendente", "Netto a Pagare", "Acconto", "Differenza", "Note"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Dati
    for row_num, salary in enumerate(salaries, 2):
        ws.cell(row=row_num, column=1, value=salary.get("dipendente_nome", ""))
        ws.cell(row=row_num, column=2, value=salary.get("netto_a_pagare", 0))
        ws.cell(row=row_num, column=3, value=salary.get("acconto_pagato", 0))
        ws.cell(row=row_num, column=4, value=salary.get("differenza", 0))
        ws.cell(row=row_num, column=5, value=salary.get("note", ""))
    
    # Larghezze colonne
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    
    # Salva
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=libro_unico_{month_year}.xlsx"}
    )


@router.put("/libro-unico/salaries/{salary_id}")
@handle_errors
async def update_libro_unico_salary(
    salary_id: str,
    netto_a_pagare: float = Query(...),
    acconto_pagato: float = Query(0),
    differenza: float = Query(...),
    note: str = Query("")
) -> Dict[str, str]:
    """Aggiorna busta paga libro unico."""
    db = Database.get_db()
    
    result = await db["libro_unico_salaries"].update_one(
        {"id": salary_id},
        {"$set": {
            "netto_a_pagare": netto_a_pagare,
            "acconto_pagato": acconto_pagato,
            "differenza": differenza,
            "note": note,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    
    return {"message": "Aggiornato"}


@router.delete("/libro-unico/salaries/{salary_id}")
@handle_errors
async def delete_libro_unico_salary(salary_id: str) -> Dict[str, str]:
    """Elimina busta paga libro unico."""
    db = Database.get_db()
    
    result = await db["libro_unico_salaries"].delete_one({"id": salary_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    
    return {"message": "Eliminato"}


# Note: salari and buste-paga routes are defined earlier in the file to avoid route conflict with /{dipendente_id}



@router.get("/portale/stats")
@handle_errors
async def get_portale_stats() -> Dict[str, Any]:
    """Statistiche portale dipendenti."""
    db = Database.get_db()
    
    total = await db[Collections.EMPLOYEES].count_documents({"attivo": {"$ne": False}})
    invitati = await db[Collections.EMPLOYEES].count_documents({"portale_invitato": True})
    registrati = await db[Collections.EMPLOYEES].count_documents({"portale_registrato": True})
    mai_invitati = await db[Collections.EMPLOYEES].count_documents({
        "attivo": {"$ne": False},
        "$or": [{"portale_invitato": False}, {"portale_invitato": {"$exists": False}}]
    })
    
    return {
        "totale": total,
        "mai_invitati": mai_invitati,
        "invitati": invitati,
        "registrati": registrati
    }


# ============== IMPORT MASSIVO LIBRETTI SANITARI ==============

@router.post("/libretti-sanitari/import-excel")
@handle_errors
async def import_libretti_sanitari_excel(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Import massivo libretti sanitari da Excel.
    Colonne richieste: Nome Dipendente, Numero Libretto, Data Rilascio, Data Scadenza, Note
    """
    import pandas as pd
    
    db = Database.get_db()
    content = await file.read()
    
    try:
        # Leggi Excel
        df = pd.read_excel(io.BytesIO(content))
        
        # Normalizza nomi colonne
        df.columns = df.columns.str.lower().str.strip().str.replace(' ', '_')
        
        created = 0
        updated = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # Estrai dati
                nome = str(row.get('nome_dipendente', row.get('dipendente', row.get('nome', '')))).strip()
                numero = str(row.get('numero_libretto', row.get('numero', ''))).strip()
                
                if not nome:
                    continue
                
                # Parse date
                data_rilascio = None
                data_scadenza = None
                
                for col in ['data_rilascio', 'rilascio', 'emissione']:
                    if col in row and pd.notna(row[col]):
                        val = row[col]
                        if isinstance(val, str):
                            for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                                try:
                                    data_rilascio = datetime.strptime(val, fmt).strftime('%Y-%m-%d')
                                    break
                                except (ValueError, TypeError):
                                    pass
                        elif hasattr(val, 'strftime'):
                            data_rilascio = val.strftime('%Y-%m-%d')
                        break
                
                for col in ['data_scadenza', 'scadenza', 'validita']:
                    if col in row and pd.notna(row[col]):
                        val = row[col]
                        if isinstance(val, str):
                            for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                                try:
                                    data_scadenza = datetime.strptime(val, fmt).strftime('%Y-%m-%d')
                                    break
                                except (ValueError, TypeError):
                                    pass
                        elif hasattr(val, 'strftime'):
                            data_scadenza = val.strftime('%Y-%m-%d')
                        break
                
                note = str(row.get('note', row.get('osservazioni', ''))).strip()
                if note == 'nan':
                    note = ''
                
                # Cerca dipendente esistente
                dipendente = await db[Collections.EMPLOYEES].find_one({
                    "$or": [
                        {"nome_completo": {"$regex": nome, "$options": "i"}},
                        {"nome": {"$regex": nome.split()[0] if nome else "", "$options": "i"}}
                    ]
                })
                
                dipendente_id = dipendente.get("id") if dipendente else None
                
                # Determina stato
                stato = "valido"
                if data_scadenza:
                    try:
                        scad = datetime.strptime(data_scadenza, '%Y-%m-%d')
                        if scad < datetime.now():
                            stato = "scaduto"
                        elif (scad - datetime.now()).days < 30:
                            stato = "in_scadenza"
                    except (ValueError, TypeError):
                        pass
                
                # Check duplicato
                existing = await db["libretti_sanitari"].find_one({
                    "$or": [
                        {"dipendente_nome": {"$regex": nome, "$options": "i"}},
                        {"numero_libretto": numero} if numero else {"_id": None}
                    ]
                })
                
                if existing:
                    # Aggiorna
                    await db["libretti_sanitari"].update_one(
                        {"id": existing["id"]},
                        {"$set": {
                            "numero_libretto": numero or existing.get("numero_libretto"),
                            "data_rilascio": data_rilascio or existing.get("data_rilascio"),
                            "data_scadenza": data_scadenza or existing.get("data_scadenza"),
                            "stato": stato,
                            "note": note or existing.get("note", ""),
                            "dipendente_id": dipendente_id or existing.get("dipendente_id"),
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )
                    updated += 1
                else:
                    # Crea nuovo
                    libretto = {
                        "id": str(uuid.uuid4()),
                        "dipendente_nome": nome,
                        "dipendente_id": dipendente_id,
                        "numero_libretto": numero,
                        "data_rilascio": data_rilascio,
                        "data_scadenza": data_scadenza,
                        "stato": stato,
                        "note": note,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db["libretti_sanitari"].insert_one(libretto.copy())
                    created += 1
                    
            except Exception as e:
                errors.append(f"Riga {idx+2}: {str(e)}")
        
        return {
            "success": True,
            "created": created,
            "updated": updated,
            "errors": errors[:20]
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Errore lettura file: {str(e)}")


@router.get("/libretti-sanitari/scadenze")
@handle_errors
async def get_libretti_in_scadenza(giorni: int = Query(30, description="Giorni per scadenza")) -> Dict[str, Any]:
    """Ritorna libretti in scadenza nei prossimi N giorni."""
    db = Database.get_db()
    
    oggi = datetime.now()
    limite = (oggi + timedelta(days=giorni)).strftime('%Y-%m-%d')
    oggi_str = oggi.strftime('%Y-%m-%d')
    
    # Scaduti
    scaduti = await db["libretti_sanitari"].find(
        {"data_scadenza": {"$lt": oggi_str}},
        {"_id": 0}
    ).sort("data_scadenza", 1).to_list(100)
    
    # In scadenza
    in_scadenza = await db["libretti_sanitari"].find(
        {"data_scadenza": {"$gte": oggi_str, "$lte": limite}},
        {"_id": 0}
    ).sort("data_scadenza", 1).to_list(100)
    
    return {
        "scaduti": scaduti,
        "in_scadenza": in_scadenza,
        "totale_scaduti": len(scaduti),
        "totale_in_scadenza": len(in_scadenza)
    }


@router.post("/libretti-sanitari/genera-da-dipendenti")
@handle_errors
async def genera_libretti_da_dipendenti() -> Dict[str, Any]:
    """
    Genera automaticamente i libretti sanitari per tutti i dipendenti attivi
    che non hanno ancora un libretto.
    """
    db = Database.get_db()
    
    # Trova dipendenti attivi
    dipendenti = await db[Collections.EMPLOYEES].find(
        {"status": {"$in": ["attivo", "active", None]}},
        {"_id": 0}
    ).to_list(500)
    
    created = 0
    skipped = 0
    
    for dip in dipendenti:
        nome = dip.get("nome_completo") or f"{dip.get('nome', '')} {dip.get('cognome', '')}".strip()
        dip_id = dip.get("id")
        
        if not nome:
            continue
        
        # Verifica se ha già libretto
        existing = await db["libretti_sanitari"].find_one({
            "$or": [
                {"dipendente_id": dip_id},
                {"dipendente_nome": {"$regex": f"^{re.escape(nome)}$", "$options": "i"}}
            ]
        })
        
        if existing:
            skipped += 1
            continue
        
        # Crea libretto vuoto
        libretto = {
            "id": str(uuid.uuid4()),
            "dipendente_nome": nome,
            "dipendente_id": dip_id,
            "numero_libretto": "",
            "data_rilascio": None,
            "data_scadenza": None,
            "stato": "da_compilare",
            "note": "Libretto generato automaticamente",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db["libretti_sanitari"].insert_one(libretto.copy())
        created += 1
    
    return {
        "success": True,
        "created": created,
        "skipped": skipped,
        "message": f"Creati {created} libretti, {skipped} dipendenti avevano già un libretto"
    }


# ============== GESTIONE CONTRATTI ==============

@router.get("/contratti")
@handle_errors
async def list_contratti(
    dipendente_id: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None),
    stato: Optional[str] = Query(None)
) -> List[Dict[str, Any]]:
    """Lista tutti i contratti."""
    db = Database.get_db()
    
    query = {}
    if dipendente_id:
        query["dipendente_id"] = dipendente_id
    if tipo:
        query["tipo_contratto"] = tipo
    if stato:
        query["stato"] = stato
    
    contratti = await db["contratti_dipendenti"].find(query, {"_id": 0}).sort("data_inizio", -1).to_list(500)
    return contratti


@router.post("/contratti")
@handle_errors
async def create_contratto(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Crea nuovo contratto per dipendente."""
    db = Database.get_db()
    
    # Verifica dipendente
    dipendente = await db[Collections.EMPLOYEES].find_one(
        {"$or": [{"id": data.get("dipendente_id")}, {"codice_fiscale": data.get("dipendente_id")}]}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")

    # Validazioni business sui dati del contratto
    retribuzione = data.get("retribuzione_lorda")
    if retribuzione is not None and float(retribuzione) < 0:
        raise HTTPException(status_code=422, detail="retribuzione_lorda non può essere negativa")

    data_inizio = data.get("data_inizio")
    data_fine = data.get("data_fine")
    if data_inizio and data_fine and str(data_fine) < str(data_inizio):
        raise HTTPException(status_code=422, detail="data_fine non può essere precedente a data_inizio")

    contratto = {
        "id": str(uuid.uuid4()),
        "dipendente_id": dipendente.get("id"),
        "dipendente_nome": dipendente.get("nome_completo") or f"{dipendente.get('nome', '')} {dipendente.get('cognome', '')}".strip(),
        "tipo_contratto": data.get("tipo_contratto", "tempo_determinato"),
        "livello": data.get("livello", ""),
        "mansione": data.get("mansione", dipendente.get("mansione", "")),
        "retribuzione_lorda": data.get("retribuzione_lorda", 0),
        "ore_settimanali": data.get("ore_settimanali", 40),
        "data_inizio": data.get("data_inizio"),
        "data_fine": data.get("data_fine"),  # None per indeterminato
        "ccnl": data.get("ccnl", "Turismo - Pubblici Esercizi"),
        "sede_lavoro": data.get("sede_lavoro", ""),
        "note": data.get("note", ""),
        "stato": "attivo",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db["contratti_dipendenti"].insert_one(contratto.copy())
    
    # Aggiorna anche il dipendente
    await db[Collections.EMPLOYEES].update_one(
        {"id": dipendente.get("id")},
        {"$set": {
            "contratto_attivo_id": contratto["id"],
            "tipo_contratto": contratto["tipo_contratto"],
            "livello": contratto["livello"],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    contratto.pop("_id", None)
    return contratto


@router.put("/contratti/{contratto_id}")
@handle_errors
async def update_contratto(contratto_id: str, data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Aggiorna contratto esistente. Non permette modifiche su contratti terminati."""
    db = Database.get_db()

    # Blocca modifiche su contratti terminati
    existing = await db["contratti_dipendenti"].find_one({"id": contratto_id}, {"_id": 0, "stato": 1})
    if not existing:
        raise HTTPException(status_code=404, detail="Contratto non trovato")
    if existing.get("stato") == "terminato":
        raise HTTPException(status_code=409, detail="Impossibile modificare un contratto terminato")
    
    data.pop("id", None)
    data.pop("created_at", None)
    data.pop("dipendente_id", None)  # Non modificabile
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db["contratti_dipendenti"].update_one(
        {"id": contratto_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contratto non trovato")
    
    return {"message": "Contratto aggiornato"}


@router.post("/contratti/{contratto_id}/termina")
@handle_errors
async def termina_contratto(
    contratto_id: str,
    data_fine: str = Query(..., description="Data fine contratto YYYY-MM-DD"),
    motivo: str = Query("", description="Motivo cessazione")
) -> Dict[str, str]:
    """Termina un contratto attivo."""
    db = Database.get_db()
    
    contratto = await db["contratti_dipendenti"].find_one({"id": contratto_id})
    if not contratto:
        raise HTTPException(status_code=404, detail="Contratto non trovato")
    
    await db["contratti_dipendenti"].update_one(
        {"id": contratto_id},
        {"$set": {
            "stato": "terminato",
            "data_fine": data_fine,
            "motivo_cessazione": motivo,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Aggiorna dipendente
    await db[Collections.EMPLOYEES].update_one(
        {"id": contratto.get("dipendente_id")},
        {"$set": {
            "contratto_attivo_id": None,
            "status": "cessato" if motivo else "attivo",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Contratto terminato"}


@router.delete("/contratti/{contratto_id}")
@handle_errors
async def delete_contratto(contratto_id: str) -> Dict[str, str]:
    """Elimina un contratto. Solo contratti in stato 'bozza' o creati per errore possono essere eliminati."""
    db = Database.get_db()
    result = await db["contratti_dipendenti"].delete_one({"id": contratto_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contratto non trovato")
    return {"message": "Contratto eliminato"}


@router.get("/contratti/scadenze")
@handle_errors
async def get_contratti_in_scadenza(giorni: int = Query(60, description="Giorni per scadenza")) -> Dict[str, Any]:
    """Ritorna contratti a tempo determinato in scadenza."""
    db = Database.get_db()
    
    oggi = datetime.now()
    limite = (oggi + timedelta(days=giorni)).strftime('%Y-%m-%d')
    oggi_str = oggi.strftime('%Y-%m-%d')
    
    # Scaduti
    scaduti = await db["contratti_dipendenti"].find(
        {
            "tipo_contratto": {"$in": ["tempo_determinato", "determinato", "Tempo Determinato"]},
            "data_fine": {"$lt": oggi_str},
            "stato": "attivo"
        },
        {"_id": 0}
    ).sort("data_fine", 1).to_list(100)
    
    # In scadenza
    in_scadenza = await db["contratti_dipendenti"].find(
        {
            "tipo_contratto": {"$in": ["tempo_determinato", "determinato", "Tempo Determinato"]},
            "data_fine": {"$gte": oggi_str, "$lte": limite},
            "stato": "attivo"
        },
        {"_id": 0}
    ).sort("data_fine", 1).to_list(100)
    
    return {
        "scaduti": scaduti,
        "in_scadenza": in_scadenza,
        "totale_scaduti": len(scaduti),
        "totale_in_scadenza": len(in_scadenza)
    }


@router.post("/contratti/import-excel")
@handle_errors
async def import_contratti_excel(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Import massivo contratti da Excel.
    Colonne: Nome Dipendente, Tipo Contratto, Livello, Mansione, Data Inizio, Data Fine, Retribuzione
    """
    import pandas as pd
    
    db = Database.get_db()
    content = await file.read()
    
    try:
        df = pd.read_excel(io.BytesIO(content))
        df.columns = df.columns.str.lower().str.strip().str.replace(' ', '_')
        
        created = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                nome = str(row.get('nome_dipendente', row.get('dipendente', row.get('nome', '')))).strip()
                if not nome or nome == 'nan':
                    continue
                
                # Cerca dipendente
                dipendente = await db[Collections.EMPLOYEES].find_one({
                    "$or": [
                        {"nome_completo": {"$regex": nome, "$options": "i"}},
                        {"nome": {"$regex": nome.split()[0] if ' ' in nome else nome, "$options": "i"}}
                    ]
                })
                
                if not dipendente:
                    errors.append(f"Riga {idx+2}: Dipendente '{nome}' non trovato")
                    continue
                
                # Parse date
                data_inizio = None
                data_fine = None
                
                for col in ['data_inizio', 'inizio', 'assunzione']:
                    if col in row and pd.notna(row[col]):
                        val = row[col]
                        if isinstance(val, str):
                            for fmt in ['%d/%m/%Y', '%Y-%m-%d']:
                                try:
                                    data_inizio = datetime.strptime(val, fmt).strftime('%Y-%m-%d')
                                    break
                                except (ValueError, TypeError):
                                    pass
                        elif hasattr(val, 'strftime'):
                            data_inizio = val.strftime('%Y-%m-%d')
                        break
                
                for col in ['data_fine', 'fine', 'scadenza']:
                    if col in row and pd.notna(row[col]):
                        val = row[col]
                        if isinstance(val, str):
                            for fmt in ['%d/%m/%Y', '%Y-%m-%d']:
                                try:
                                    data_fine = datetime.strptime(val, fmt).strftime('%Y-%m-%d')
                                    break
                                except (ValueError, TypeError):
                                    pass
                        elif hasattr(val, 'strftime'):
                            data_fine = val.strftime('%Y-%m-%d')
                        break
                
                tipo = str(row.get('tipo_contratto', row.get('tipo', 'tempo_determinato'))).strip()
                if tipo == 'nan':
                    tipo = 'tempo_determinato'
                
                retribuzione = 0
                for col in ['retribuzione', 'retribuzione_lorda', 'stipendio', 'ral']:
                    if col in row and pd.notna(row[col]):
                        try:
                            retribuzione = float(row[col])
                        except (ValueError, TypeError):
                            pass
                        break
                
                contratto = {
                    "id": str(uuid.uuid4()),
                    "dipendente_id": dipendente.get("id"),
                    "dipendente_nome": dipendente.get("nome_completo") or nome,
                    "tipo_contratto": tipo,
                    "livello": str(row.get('livello', '')).strip() if pd.notna(row.get('livello')) else "",
                    "mansione": str(row.get('mansione', dipendente.get('mansione', ''))).strip() if pd.notna(row.get('mansione')) else dipendente.get('mansione', ''),
                    "retribuzione_lorda": retribuzione,
                    "ore_settimanali": int(row.get('ore', row.get('ore_settimanali', 40))) if pd.notna(row.get('ore', row.get('ore_settimanali'))) else 40,
                    "data_inizio": data_inizio,
                    "data_fine": data_fine,
                    "ccnl": str(row.get('ccnl', 'Turismo - Pubblici Esercizi')).strip() if pd.notna(row.get('ccnl')) else 'Turismo - Pubblici Esercizi',
                    "stato": "attivo",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db["contratti_dipendenti"].insert_one(contratto.copy())
                
                # Aggiorna dipendente
                await db[Collections.EMPLOYEES].update_one(
                    {"id": dipendente.get("id")},
                    {"$set": {
                        "contratto_attivo_id": contratto["id"],
                        "tipo_contratto": tipo,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                
                created += 1
                
            except Exception as e:
                errors.append(f"Riga {idx+2}: {str(e)}")
        
        return {
            "success": True,
            "created": created,
            "errors": errors[:20]
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Errore lettura file: {str(e)}")



# ============== IMPORT BUSTE PAGA ==============

@router.get("/buste-paga/scan")
@handle_errors
async def scan_buste_paga_folders() -> Dict[str, Any]:
    """
    Scansiona le cartelle delle buste paga e restituisce i progressivi trovati.
    """
    import os
    from app.utils.busta_paga_parser import scan_all_dipendenti
    
    base_path = "/app/documents/buste_paga"
    
    if not os.path.exists(base_path):
        return {
            "success": False,
            "error": "Cartella buste paga non trovata",
            "path": base_path
        }
    
    # Scansiona tutte le cartelle
    progressivi = scan_all_dipendenti(base_path)
    
    # Lista cartelle disponibili
    cartelle = [d for d in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, d))]
    
    return {
        "success": True,
        "cartelle_trovate": len(cartelle),
        "dipendenti_con_progressivi": len(progressivi),
        "cartelle": cartelle,
        "progressivi": progressivi
    }


@router.post("/buste-paga/import")
@handle_errors
async def import_buste_paga_to_dipendenti(
    match_mode: str = Query("cf", description="Modalità matching: cf (codice fiscale), nome (nome completo)"),
    dry_run: bool = Query(True, description="Se true, mostra solo le corrispondenze senza aggiornare")
) -> Dict[str, Any]:
    """
    Importa i progressivi dalle buste paga nei record dei dipendenti.
    
    - match_mode: "cf" per codice fiscale, "nome" per nome completo
    - dry_run: se True, mostra solo le corrispondenze trovate senza aggiornare
    """
    from app.utils.busta_paga_parser import scan_all_dipendenti
    
    db = Database.get_db()
    base_path = "/app/documents/buste_paga"
    
    # Carica tutti i dipendenti dal DB
    dipendenti_db = await db[Collections.EMPLOYEES].find({}, {"_id": 0}).to_list(1000)
    
    # Scansiona le buste paga
    progressivi_bp = scan_all_dipendenti(base_path)
    
    matches = []
    no_matches = []
    updates = []
    errors = []
    
    for folder_name, progressivi in progressivi_bp.items():
        nome_normalizzato = folder_name.lower().replace('_', ' ')
        matched = False
        
        for dip in dipendenti_db:
            dip_nome = (dip.get('nome_completo') or f"{dip.get('nome', '')} {dip.get('cognome', '')}").lower().strip()
            dip_cf = (dip.get('codice_fiscale') or '').upper()
            
            # Match per nome
            nome_match = False
            if match_mode == "nome":
                # Prova match esatto o parziale
                nome_match = (nome_normalizzato == dip_nome or 
                             nome_normalizzato in dip_nome or 
                             dip_nome in nome_normalizzato)
                
                # Prova anche invertendo nome/cognome
                if not nome_match:
                    parts = nome_normalizzato.split()
                    if len(parts) >= 2:
                        inverted = f"{parts[-1]} {' '.join(parts[:-1])}"
                        nome_match = inverted == dip_nome or inverted in dip_nome
            
            # Match per CF se disponibile
            cf_match = False
            if match_mode == "cf" and progressivi.get('codice_fiscale'):
                cf_match = progressivi.get('codice_fiscale', '').upper() == dip_cf
            
            if nome_match or cf_match:
                matched = True
                match_info = {
                    "cartella": folder_name,
                    "dipendente_db": dip.get('nome_completo') or f"{dip.get('nome', '')} {dip.get('cognome', '')}",
                    "dipendente_id": dip.get('id'),
                    "match_type": "cf" if cf_match else "nome",
                    "progressivi": progressivi
                }
                matches.append(match_info)
                
                if not dry_run:
                    # Aggiorna il dipendente con i progressivi
                    try:
                        update_data = {
                            "paga_base": progressivi.get('paga_base', 0),
                            "contingenza": progressivi.get('contingenza', 0),
                            "progressivi": {
                                "tfr_accantonato": progressivi.get('tfr_accantonato', 0),
                                "tfr_quota_anno": progressivi.get('tfr_quota_anno', 0),
                                "ferie_maturate": progressivi.get('ferie_maturate', 0),
                                "ferie_godute": progressivi.get('ferie_godute', 0),
                                "ferie_residue": progressivi.get('ferie_residue', 0),
                                "permessi_maturati": progressivi.get('permessi_maturati', 0),
                                "permessi_goduti": progressivi.get('permessi_goduti', 0),
                                "permessi_residui": progressivi.get('permessi_residui', 0),
                                "rol_maturati": progressivi.get('rol_maturati', 0),
                                "rol_goduti": progressivi.get('rol_goduti', 0),
                                "rol_residui": progressivi.get('rol_residui', 0),
                                "anno_riferimento": progressivi.get('anno_riferimento'),
                                "mese_riferimento": progressivi.get('mese_riferimento'),
                                "fonte_busta_paga": progressivi.get('fonte')
                            },
                            "updated_at": datetime.now(timezone.utc).isoformat(),
                            "progressivi_importati_at": datetime.now(timezone.utc).isoformat()
                        }
                        
                        await db[Collections.EMPLOYEES].update_one(
                            {"id": dip.get('id')},
                            {"$set": update_data}
                        )
                        updates.append(dip.get('id'))
                    except Exception as e:
                        errors.append({"id": dip.get('id'), "error": str(e)})
                
                break
        
        if not matched:
            no_matches.append(folder_name)
    
    return {
        "success": True,
        "dry_run": dry_run,
        "match_mode": match_mode,
        "totale_cartelle": len(progressivi_bp),
        "matches_trovati": len(matches),
        "non_trovati": len(no_matches),
        "aggiornati": len(updates) if not dry_run else 0,
        "errori": len(errors),
        "matches": matches,
        "no_matches": no_matches,
        "errors": errors if errors else None
    }


@router.get("/buste-paga/dipendente/{dipendente_id}")
@handle_errors
async def get_buste_paga_dipendente(dipendente_id: str) -> Dict[str, Any]:
    """
    Restituisce tutte le buste paga trovate per un dipendente specifico.
    Cerca per nome o CF nelle cartelle delle buste paga.
    """
    import os
    from app.utils.busta_paga_parser import scan_dipendente_folder
    
    db = Database.get_db()
    
    # Trova il dipendente nel DB
    dipendente = await db[Collections.EMPLOYEES].find_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"_id": 0}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    nome_completo = dipendente.get('nome_completo') or f"{dipendente.get('nome', '')} {dipendente.get('cognome', '')}"
    
    # Cerca la cartella corrispondente
    base_path = "/app/documents/buste_paga"
    cartelle = os.listdir(base_path) if os.path.exists(base_path) else []
    
    cartella_trovata = None
    for cartella in cartelle:
        nome_cartella = cartella.lower().replace('_', ' ')
        if nome_completo.lower() in nome_cartella or nome_cartella in nome_completo.lower():
            cartella_trovata = cartella
            break
        # Prova anche invertendo
        parts = nome_completo.lower().split()
        if len(parts) >= 2:
            inverted = f"{parts[-1]} {' '.join(parts[:-1])}"
            if inverted in nome_cartella or nome_cartella in inverted:
                cartella_trovata = cartella
                break
    
    if not cartella_trovata:
        return {
            "success": False,
            "dipendente": nome_completo,
            "message": "Cartella buste paga non trovata per questo dipendente",
            "cartelle_disponibili": cartelle[:20]
        }
    
    # Scansiona la cartella
    folder_path = os.path.join(base_path, cartella_trovata)
    buste = scan_dipendente_folder(folder_path)
    
    return {
        "success": True,
        "dipendente": nome_completo,
        "cartella": cartella_trovata,
        "totale_buste": len(buste),
        "buste": buste
    }



@router.post("/buste-paga/dipendente/{dipendente_id}/import")
@handle_errors
async def import_busta_paga_to_dipendente(dipendente_id: str) -> Dict[str, Any]:
    """
    Importa i progressivi dall'ultima busta paga trovata per un dipendente specifico.
    Aggiorna il record del dipendente con TFR, ferie, permessi, paga base, contingenza.
    """
    import os
    from app.utils.busta_paga_parser import get_latest_progressivi
    
    db = Database.get_db()
    
    # Trova il dipendente nel DB
    dipendente = await db[Collections.EMPLOYEES].find_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"_id": 0}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    nome_completo = dipendente.get('nome_completo') or f"{dipendente.get('nome', '')} {dipendente.get('cognome', '')}"
    
    # Cerca la cartella corrispondente
    base_path = "/app/documents/buste_paga"
    cartelle = os.listdir(base_path) if os.path.exists(base_path) else []
    
    cartella_trovata = None
    for cartella in cartelle:
        nome_cartella = cartella.lower().replace('_', ' ')
        if nome_completo.lower() in nome_cartella or nome_cartella in nome_completo.lower():
            cartella_trovata = cartella
            break
        # Prova anche invertendo
        parts = nome_completo.lower().split()
        if len(parts) >= 2:
            inverted = f"{parts[-1]} {' '.join(parts[:-1])}"
            if inverted in nome_cartella or nome_cartella in inverted:
                cartella_trovata = cartella
                break
    
    if not cartella_trovata:
        return {
            "success": False,
            "dipendente": nome_completo,
            "message": "Cartella buste paga non trovata per questo dipendente"
        }
    
    # Ottieni i progressivi
    folder_path = os.path.join(base_path, cartella_trovata)
    progressivi = get_latest_progressivi(folder_path)
    
    if not progressivi:
        return {
            "success": False,
            "dipendente": nome_completo,
            "cartella": cartella_trovata,
            "message": "Nessun progressivo trovato nelle buste paga"
        }
    
    # Aggiorna il dipendente
    update_data = {
        "paga_base": progressivi.get('paga_base', 0),
        "contingenza": progressivi.get('contingenza', 0),
        "progressivi": {
            "tfr_accantonato": progressivi.get('tfr_accantonato', 0),
            "tfr_quota_anno": progressivi.get('tfr_quota_anno', 0),
            "ferie_maturate": progressivi.get('ferie_maturate', 0),
            "ferie_godute": progressivi.get('ferie_godute', 0),
            "ferie_residue": progressivi.get('ferie_residue', 0),
            "permessi_maturati": progressivi.get('permessi_maturati', 0),
            "permessi_goduti": progressivi.get('permessi_goduti', 0),
            "permessi_residui": progressivi.get('permessi_residui', 0),
            "rol_maturati": progressivi.get('rol_maturati', 0),
            "rol_goduti": progressivi.get('rol_goduti', 0),
            "rol_residui": progressivi.get('rol_residui', 0),
            "anno_riferimento": progressivi.get('anno_riferimento'),
            "mese_riferimento": progressivi.get('mese_riferimento'),
            "fonte_busta_paga": progressivi.get('fonte')
        },
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "progressivi_importati_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db[Collections.EMPLOYEES].update_one(
        {"id": dipendente.get('id')},
        {"$set": update_data}
    )
    
    return {
        "success": True,
        "dipendente": nome_completo,
        "cartella": cartella_trovata,
        "progressivi_importati": update_data,
        "fonte": progressivi.get('fonte')
    }



# ============== REPORT PDF FERIE/PERMESSI ==============

@router.get("/{dipendente_id}/report-ferie-permessi")
@handle_errors
async def genera_report_ferie_permessi(
    dipendente_id: str,
    anno: int = Query(None, description="Anno di riferimento (default: anno corrente)")
) -> Dict[str, Any]:
    """
    Genera report PDF annuale ferie e permessi per un dipendente.
    Include: progressivi, storico mensile, riepilogo annuale.
    """
    from fastapi.responses import Response
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER
    import io
    
    db = Database.get_db()
    
    if not anno:
        anno = datetime.now().year
    
    # Trova dipendente
    dipendente = await db[Collections.EMPLOYEES].find_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"_id": 0}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    nome = dipendente.get("nome_completo") or dipendente.get("name") or f"{dipendente.get('cognome', '')} {dipendente.get('nome', '')}"
    cf = dipendente.get("codice_fiscale", "N/A")
    progressivi = dipendente.get("progressivi", {})
    
    # Recupera cedolini dell'anno per storico mensile
    cedolini = await db["cedolini"].find({
        "$or": [
            {"dipendente_id": dipendente_id},
            {"codice_fiscale": cf}
        ],
        "anno": anno
    }, {"_id": 0}).sort("mese", 1).to_list(12)
    
    # Prepara dati
    ferie_maturate = progressivi.get("ferie_maturate", 0)
    ferie_godute = progressivi.get("ferie_godute", 0)
    ferie_residue = progressivi.get("ferie_residue", 0)
    permessi_maturati = progressivi.get("permessi_maturati", 0)
    permessi_goduti = progressivi.get("permessi_goduti", 0)
    permessi_residui = progressivi.get("permessi_residui", 0)
    rol_maturati = progressivi.get("rol_maturati", 0)
    rol_goduti = progressivi.get("rol_goduti", 0)
    rol_residui = progressivi.get("rol_residui", 0)
    
    # Genera PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=20)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, alignment=TA_CENTER, spaceAfter=10)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=10)
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold')
    
    elements = []
    
    # Intestazione
    elements.append(Paragraph(f"REPORT FERIE E PERMESSI {anno}", title_style))
    elements.append(Paragraph(f"{nome}", subtitle_style))
    elements.append(Paragraph(f"Codice Fiscale: {cf}", normal_style))
    elements.append(Paragraph(f"Data generazione: {datetime.now().strftime('%d/%m/%Y %H:%M')}", normal_style))
    elements.append(Spacer(1, 20))
    
    # Tabella Riepilogo
    elements.append(Paragraph("RIEPILOGO PROGRESSIVI", header_style))
    elements.append(Spacer(1, 10))
    
    data_riepilogo = [
        ["Voce", "Maturate", "Godute", "Residue"],
        ["Ferie (giorni)", f"{ferie_maturate:.1f}", f"{ferie_godute:.1f}", f"{ferie_residue:.2f}"],
        ["Permessi (ore)", f"{permessi_maturati:.2f}", f"{permessi_goduti:.2f}", f"{permessi_residui:.2f}"],
        ["ROL (ore)", f"{rol_maturati:.2f}", f"{rol_goduti:.2f}", f"{rol_residui:.2f}"],
    ]
    
    table_riepilogo = Table(data_riepilogo, colWidths=[6*cm, 3*cm, 3*cm, 3*cm])
    table_riepilogo.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(table_riepilogo)
    elements.append(Spacer(1, 30))
    
    # Tabella Storico Mensile (se ci sono cedolini)
    if cedolini:
        elements.append(Paragraph("DETTAGLIO MENSILE", header_style))
        elements.append(Spacer(1, 10))
        
        mesi_nomi = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]
        
        data_mensile = [["Mese", "Ferie Godute", "Permessi Goduti", "Netto Busta"]]
        
        for ced in cedolini:
            mese_num = ced.get("mese", 0)
            mese_nome = mesi_nomi[mese_num - 1] if 1 <= mese_num <= 12 else str(mese_num)
            ferie_g = ced.get("ferie_godute", 0) or 0
            perm_g = ced.get("permessi_goduti", ced.get("ore_permesso", 0)) or 0
            netto = ced.get("netto", ced.get("netto_mese", 0)) or 0
            
            data_mensile.append([
                f"{mese_nome} {anno}",
                f"{ferie_g:.1f}" if ferie_g else "-",
                f"{perm_g:.1f}" if perm_g else "-",
                f"€ {netto:,.2f}"
            ])
        
        table_mensile = Table(data_mensile, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
        table_mensile.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(table_mensile)
        elements.append(Spacer(1, 20))
    
    # Note finali
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("NOTE", header_style))
    elements.append(Paragraph(f"• Anno di riferimento progressivi: {progressivi.get('anno_riferimento', anno)}", normal_style))
    elements.append(Paragraph(f"• Mese di riferimento: {progressivi.get('mese_riferimento', 'N/A')}", normal_style))
    if progressivi.get('fonte_busta_paga'):
        elements.append(Paragraph(f"• Fonte dati: {progressivi.get('fonte_busta_paga')}", normal_style))
    
    # Footer
    elements.append(Spacer(1, 40))
    elements.append(Paragraph("_" * 50, normal_style))
    elements.append(Paragraph("Documento generato automaticamente dal sistema ERP", 
                              ParagraphStyle('Footer', parent=normal_style, fontSize=8, textColor=colors.gray)))
    
    # Build PDF
    doc.build(elements)
    
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    filename = f"Report_Ferie_Permessi_{nome.replace(' ', '_')}_{anno}.pdf"
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )





